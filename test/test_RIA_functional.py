import json
from MpApi.Utils.Ria import init_ria, RIA, record_exists, record_exists2
from openpyxl import Workbook, load_workbook, worksheet
from pathlib import Path
# from openpyxl.cell.cell import Cell
# from openpyxl.styles.colors import Color


def test_init_ria() -> None:
    client = init_ria()
    assert isinstance(client, RIA)


def _record_exists(Type: str) -> None:
    excel_fn = (
        Path(__file__).parents[1]
        / "sdata"
        / "Abschrift_HK_Afrika_III_C_Testimport_Erw_Art_bereinigt_.xlsx"
    )
    wb = load_workbook(excel_fn, data_only=True)
    # ws = wb["Erw. Art bereinigt"]
    ws = wb.active
    conf = {}
    conf["org_unit"] = "EMAfrika1"
    conf["RIA"] = init_ria()

    results = {}
    r: bool = False
    for (
        idx,
        row,
    ) in enumerate(ws.iter_rows(min_row=2), start=2):
        ident = row[0].value
        if Type == "exists" and ident is not None:
            r = record_exists(ident=ident, conf=conf)
        elif Type == "exists2" and ident is not None:
            r = record_exists2(ident=ident, conf=conf)

        print(f"{idx}: {ident}: {r}")
        results[ident] = r
    assert 1 == 1
    print("Writing results to {Type}.json")
    with open(f"{Type}.json", "w") as f:
        json.dump(results, f, indent=4)


def yytest_record_exists() -> None:
    _record_exists(Type="exists")


def test_record_exists2() -> None:
    _record_exists(Type="exists2")
