# from mpapi.constants import get_credentials
# from mpapi.search import Search
from copy import deepcopy
from MpApi.Utils.becky.cache_ops import open_person_cache, save_person_cache
from MpApi.Utils.becky.set_fields_Object import (
    _get_name_date,
    _lookup_name,
    _quad_split,
    _sanitize,
    _sanitize_multi,
    _split_off_prefix,
    _split_off_role,
    roles,
    set_ident,
    set_ident_sort,
    set_sachbegriff,
    set_beteiligte,
    set_erwerbDatum,
    set_erwerbungsart,
    set_erwerbNr,
    set_erwerbVon,
    set_geogrBezug,
    set_invNotiz,
    set_objRefA,
)
from MpApi.Utils.Ria import RIA, init_ria
from pathlib import Path
import pytest

# conf_fn = Path(__file__).parents[1] / "sdata" / "becky_conf.toml"


def test_lookup_name() -> None:
    """
    test this later
    _lookup_name looks up if a name exists in the cache and typically returns an int
    """
    conf = {"project_dir": Path("../sdata"), "person_cache": "person_cache.toml"}
    cases = {"Bruno von Rauchhaupt": 3269}
    for name in cases:
        pkId = _lookup_name(name=name, conf=conf)
        assert pkId == cases[name]

    # doesnt exist, should raise
    with pytest.raises(KeyError):
        pkId = _lookup_name(name="doesnt exist", conf=conf)

    # Serdu has no pkId in cache at the moment, see Serdu (?)
    with pytest.raises(KeyError):
        pkId = _lookup_name(name="Serdu", conf=conf)


def test_lookup_person() -> None:
    conf = {
        "person_cache": "person_cache.toml",
        "project_dir": Path(__file__).parents[1] / "sdata",
    }

    # bad test. It depends on the data currently in the person_cache.toml
    valid_cases = {
        "Ferdinand Werne": 3538,
        "Christian Gottfried Ehrenberg": 2219,
    }

    # why is this person problematic? Because he used to exist twice in index?
    # problematic = {"Carl Ritter": [2438]}

    for name in valid_cases:
        assert _lookup_name(name=name, conf=conf) == valid_cases[name]

    # for name in problematic:
    #    assert _lookup_name(name=name, conf=conf) == problematic[name]
    # doesn't raise anymore
    # for name in problematic:
    #    with pytest.raises(TypeError):
    #        _lookup_name(name=name, conf=conf)


#
# live tests requiring access to RIA
#


def test_set_ident() -> None:
    """
    TODO: I should be testing all of the fields that are set
    - InventarNrSTxt,
    - Part1Txt,
    - Part2Txt,
    - Part3Txt,
    - Part4Txt,
    - SortLnu,
    - DenominationVoc,
    - InvNumberSchemeRef
    """

    client = init_ria()
    templateM = client.get_template(ID=625690, mtype="Object")
    recordM = deepcopy(templateM)  # record should contain only one moduleItem
    set_ident(recordM, ident="III C 123", institution="EM")
    # print(recordM)
    # recordM.toFile(path="test.debug.xml")
    InventarNrSTxt = recordM.xpath("""/m:application/m:modules/m:module[
        @name = 'Object'
    ]/m:moduleItem/m:repeatableGroup[
        @name = 'ObjObjectNumberGrp'
    ]/m:repeatableGroupItem/m:dataField[@name ='InventarNrSTxt']/m:value/text()""")[0]

    assert InventarNrSTxt == "III C 123"


def test_set_beteiligte() -> None:
    # test doesn't work yet in a meaningful way
    conf = {
        "person_cache": "person_cache.toml",
        "project_dir": Path(__file__).parents[1] / "sdata",
    }

    client = init_ria()
    templateM = client.get_template(ID=625690, mtype="Object")
    recordM = deepcopy(templateM)  # record should contain only one moduleItem
    # bad test. Depends on data in current person_cache.toml
    set_beteiligte(recordM, beteiligte="Ferdinand Werne", conf=conf, missing_info=False)
    # print(recordM)
    # recordM.toFile(path="test.debug.xml")
    InventarNrSTxt = recordM.xpath("""/m:application/m:modules/m:module[
        @name = 'Object'
    ]/m:moduleItem/m:repeatableGroup[
        @name = 'ObjObjectNumberGrp'
    ]/m:repeatableGroupItem/m:dataField[
        @name ='InventarNrSTxt'
    ]/m:value/text()""")[0]


def test_set_erwerbdatum() -> None:
    # ObjAcquisitionDateGrp
    conf = {
        "person_cache": "person_cache.toml",
        "project_dir": Path(__file__).parents[1] / "sdata",
    }

    client = init_ria()
    templateM = client.get_template(ID=625690, mtype="Object")
    recordM = deepcopy(templateM)  # record should contain only one moduleItem
    set_erwerbDatum(recordM, datum="1.1.2100")

    # todo: check if recordM has the desired information
    xpath = """
        /m:application/m:modules/m:module[
            @name = 'Object'
        ]/m:moduleItem/m:repeatableGroup[
            @name = 'ObjAcquisitionDateGrp'
        ]/m:repeatableGroupItem/m:dataField[
            @name ='DateToTxt'
        ]/m:value/text()
    """
    assert recordM.xpath(xpath)[0] == "1.1.2100"
    set_erwerbDatum(recordM, datum="2.1.2100")
    assert recordM.xpath(xpath)[0] == "2.1.2100"


def test_sanitize() -> None:
    cases = {  # working cases
        "AKG & Co": "AKG &amp; Co",
    }
    for case in cases:
        res = _sanitize(case)
        assert res == cases[case]

    # raises
    with pytest.raises(TypeError):
        res = _sanitize(None)

    cases = ["", " ", " \n "]
    with pytest.raises(ValueError):
        for case in cases:
            res = _sanitize(case)


def test_sanitize_multi() -> None:
    cases = {"Bafo": 1, "Bafo; Baffalo": 2, "Bafo;": 1}
    for case in cases:
        alist = _sanitize_multi(case)
        assert len(alist) == cases[case]


def test_triple_split_multi() -> None:
    beteiligte = """
        Heinrich Barth (16.2.1821 - 25.11.1865), Sammler*in; 
        Königliche Preußische Kunstkammer, Ethnografische Abteilung (1801 - 1873), Vorbesitzer*in
    """
    beteiligteL = _sanitize_multi(beteiligte)
    for idx, beteiligte2 in enumerate(beteiligteL):
        prefix, name, date, role = _quad_split(beteiligte2)
        match idx:
            case 0:
                assert name == "Heinrich Barth"
                assert role == "Sammler*in"
                assert date == "16.2.1821 - 25.11.1865"
            case 1:
                assert (
                    name
                    == "Königliche Preußische Kunstkammer, Ethnografische Abteilung"
                )
                assert role == "Vorbesitzer*in"
                assert date == "1801 - 1873"


def test_split_off_role() -> None:
    cases = [
        "Claus Schilling (5.7.1871 (?) - 1946), Sammler*in",
        "Claus Schilling (5.7.1871 (?) - 1946)",
        "",
        "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan), Veräußerung",
    ]

    for idx, case in enumerate(cases):
        left, role = _split_off_role(case)
        match idx:
            case 0:
                assert left == "Claus Schilling (5.7.1871 (?) - 1946)"
                assert role == "Sammler*in"
            case 1:
                assert left == "Claus Schilling (5.7.1871 (?) - 1946)"
                assert role is None
            case 2:
                assert left is None
                assert role is None
            case 3:
                assert (
                    left
                    == "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)"
                )
                assert role == "Veräußerung"


def test_split_off_prefix() -> None:
    cases = [
        "Claus Schilling (5.7.1871 (?) - 1946), Sammler*in",
        "Prefix: Claus Schilling (5.7.1871 (?) - 1946)",
        "",
    ]

    for idx, case in enumerate(cases):
        prefix, right = _split_off_prefix(case)
        match idx:
            case 0:
                assert prefix is None
                assert right == "Claus Schilling (5.7.1871 (?) - 1946), Sammler*in"
            case 1:
                assert prefix == "Prefix"
                assert right == "Claus Schilling (5.7.1871 (?) - 1946)"
            case 2:
                assert prefix is None
                assert right is None


def test_name_date() -> None:
    cases = [
        "Claus Schilling (5.7.1871 (?) - 1946)",
        "Claus Schilling (geb. Schiller) (5.7.1871 (?) - 1946)",
        "Claus Schilling",
        "",
        "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)",  # no date
        "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)(1962)",  # no date
    ]

    for idx, case in enumerate(cases):
        name, date = _get_name_date(case)
        match idx:
            case 0:
                assert name == "Claus Schilling"
                assert date == "5.7.1871 (?) - 1946"
            case 1:
                assert name == "Claus Schilling (geb. Schiller)"
                assert date == "5.7.1871 (?) - 1946"
            case 2:
                assert name == "Claus Schilling"
                assert date is None
            case 3:
                assert name is None
                assert date is None
            case 4:
                assert (
                    name
                    == "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)"
                )
                assert date is None
            case 5:
                assert (
                    name
                    == "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)"
                )
                assert date == "1962"


def test_quad_split() -> None:
    cases = [
        "Claus Schilling (5.7.1871 (?) - 1946), Sammler*in",
        "Joachim Pfeil (30.12.1857 - 12.3.1924), Sammler*in",
        "Kaiserliches Auswärtiges Amt des Deutschen Reiches (1875), Veräußerung",
        "Bezug unklar: Paul Grade († 05.04.1894*)",
        "A. Palamidessi (?) (1939), Veräußerung",
        "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)(1962), Veräußerung",
        "Alex(ander) Siebold (1872), Veräußerung",
        'Augusta Kell ("Gulla") Pfeffer (1887 - 1967), Veräußerung',
        "Baboo Mukharji (Muckharjee) (1892), Veräußerung",
        "British Museum (Dep. of Ethnography) (1893), Veräußerung",
    ]

    # "Brunhilde Körner (geb. Lessing) (01.04.1949 - 30.04.1973), Sammler*in": {
    # "name": "Brunhilde Körner (geb. Lessing)",
    # "date": "01.04.1949 - 30.04.1973",
    # "role": "Sammler*in",
    # },
    # "Brunhilde Körner (geb. Lessing), Sammler*in": {
    # "name": "Brunhilde Körner (geb. Lessing)",
    # "date": "01.04.1949 - 30.04.1973",
    # "role": "Sammler*in",
    # },
    # "China-Bohlken (Theodor Bohlken) (1922)": {
    # "name": "China-Bohlken (Theodor Bohlken)",
    # "date": "1922",
    # "role": None,
    # },
    # "China-Bohlken (Theodor Bohlken) (1922), Veräußerung": {
    # "name": "China-Bohlken (Theodor Bohlken)",
    # "date": "1922",
    # "role": "Veräußerung",
    # },
    # "China-Bohlken (Theodor Bohlken), Veräußerung": {
    # "name": "China-Bohlken (Theodor Bohlken)",
    # "date": "1922",
    # "role": "Veräußerung",
    # },
    # "China-Bohlken (Theodor Bohlken), Sammler*in": {
    # "name": "China-Bohlken (Theodor Bohlken)",
    # "date": "1922",
    # "role": "Veräußerung",
    # },
    # "Chinesisch-Tibetische Delegation (I), Vorbesitzer*in": {
    # "name": "Chinesisch-Tibetische Delegation (I)",
    # "date": "1954",
    # "role": "Vorbesitzer*in",
    # },
    # "Chinesisch-Tibetische Delegation (II), Veräußerung": {
    # "name": "Chinesisch-Tibetische Delegation (II)",
    # "date": "1991",
    # "role": "Veräußerung",
    # },
    # "Chou I-Hsiung 周義雄 (Yü-Heng), Vorbesitzer*in": {
    # "name": "Chou I-Hsiung 周義雄 (Yü-Heng)",
    # "date": "1943",
    # "role": "Vorbesitzer*in",
    # },
    # "Claus Schilling (5.7.1871 (?) - 1946), Sammler*in": {
    # "name": "Claus Schilling",
    # "date": "5.7.1871 (?) - 1946",
    # "role": "Sammler*in",
    # },
    # "Deutsche Armee-, Marine- und Kolonialausstellung (D.A.M.U.K.A.) (1907), Vorbesitzer*in": {
    # "name": "Deutsche Armee-, Marine- und Kolonialausstellung (D.A.M.U.K.A.)",
    # "date": "1907",
    # "role": "Vorbesitzer*in",
    # },
    # "Duan-fang (Tuan-Fang) 端方 (1906), Veräußerung": {
    # "name": "Duan-fang (Tuan-Fang) 端方",
    # "date": "1906",
    # "role": "Veräußerung",
    # },
    # "Duan-fang (Tuan-Fang) 端方, Veräußerung": {
    # "name": "Duan-fang (Tuan-Fang) 端方",
    # "date": "1906",
    # "role": "Veräußerung",
    # },
    # "Enrico Hillyer (Henry) Giglioli (1845 - 1909), Veräußerung": {
    # "name": "Enrico Hillyer (Henry) Giglioli",
    # "date": "1845 - 1909",
    # "role": "Veräußerung",
    # },
    # "Fernsehteam der Nippon Hoso Kyokai (Japanische Rundfunkgesellschaft (NHK), Sammler*in": {
    # "name": "Fernsehteam der Nippon Hoso Kyokai (Japanische Rundfunkgesellschaft (NHK)",
    # "date": "1996",
    # "role": "Sammler*in",
    # },
    # "Firma Ernst Fritzsche (China-Fritzsche), Veräußerung": {
    # "name": "Firma Ernst Fritzsche (China-Fritzsche)",
    # "date": "1963",
    # "role": "Veräußerung",
    # },
    # "Forschungsreise Prof. Bernhard Struck & Dr. Hugo Bernatzik (1930 - 1931), Sammler*in": {
    # "name": "Forschungsreise Prof. Bernhard Struck & Dr. Hugo Bernatzik",
    # "date": "1930 - 1931",
    # "role": "Sammler*in",
    # },
    # "Firma Ernst Fritzsche (China-Fritzsche) (1963), Veräußerung": {
    # "name": "Firma Ernst Fritzsche (China-Fritzsche)",
    # "date": "1963",
    # "role": "Veräußerung",
    # },
    # "Frau Stange (geb. Dominik) (1956), Veräußerung": {
    # "name": "Frau Stange (geb. Dominik)",
    # "date": "1956",
    # "role": "Veräußerung",
    # },
    # "Futabashi (Nihashi) Kadaiko (?) 二橋加代子, Hersteller*in": {
    # "name": "Futabashi (Nihashi) Kadaiko (?) 二橋加代子",
    # "date": "1931",
    # "role": "Hersteller*in",
    # },
    # "Frau Pawel (Pavel), Veräußerung": {
    # "name": "Frau Pawel (Pavel)",
    # "date": "1902",
    # "role": "Veräußerung",
    # },
    # "Galerie Carrefour (M. Pierre Vérité) (1964/1956), Verbesitzer*in": {
    # "name": "Galerie Carrefour (M. Pierre Vérité)",
    # "date": "1964/1956",
    # "role": "Vorbesitzer*in",
    # },
    # "Galerie Carrefour (M. Pierre Vérité) (1964/1956), Veräußerung": {
    # "name": "Galerie Carrefour (M. Pierre Vérité)",
    # "date": "1964/1956",
    # "role": "Veräußerung",
    # },
    # "Guo Ruping 郭女屏 (Ostasiatika-Kuo); Veräußerung": {
    # "name": "Guo Ruping 郭女屏 (Ostasiatika-Kuo)",
    # "date": "1971",
    # "role": "Veräußerung",
    # },
    # "Heinrich Peters (Ostasiatische Kunst), Vorbesitzer*in": {
    # "name": "Heinrich Peters (Ostasiatische Kunst)",
    # "date": "1953",
    # "role": "Vorbesitzer*in",
    # },
    # "Heinrich Peters (Ostasiatische Kunst) (1953), Veräußerung": {
    # "name": "Heinrich Peters (Ostasiatische Kunst)",
    # "date": "1953",
    # "role": "Veräußerung",
    # },
    # "Heinrich Peters (Ostasiatische Kunst), Veräußerung": {
    # "name": "Heinrich Peters (Ostasiatische Kunst)",
    # "date": "1953",
    # "role": "Veräußerung",
    # },
    # "Helmut SchmI Dt (1982), Veräußerung": {
    # "name": "Helmut Schmidt",
    # "date": "1982",
    # "role": "Veräußerung",
    # },
    # "Herbert Credé (Kunst & Antiquitäten), Sammler*in": {
    # "name": "Herbert Credé (Kunst & Antiquitäten)",
    # "date": "1960",
    # "role": "Sammler*in",
    # },
    # "Herbert Credé (Kunst & Antiquitäten), Veräußerung": {
    # "name": "Herbert Credé (Kunst & Antiquitäten)",
    # "date": "1960",
    # "role": "Veräußerung",
    # },
    # "Idrissou (Majesté) Njoya (2020), Maler*in des Originals": {
    # "name": "Idrissou (Majesté) Njoya",
    # "date": "2020",
    # "role": "Maler*in des Originals",
    # },
    # "Ignaz (Ignatius) Sichelbarth": {
    # "name": "Ignaz (Ignatius) Sichelbarth",
    # "date": "26.9.1708 - 6.10.1780",
    # "role": None,
    # },
    # "Internationale Handwerksausstellung (1938) (28.5.1938 - 10.7.1938), Veräußerung": {
    # "name": "Internationale Handwerksausstellung",
    # "date": "1938",
    # "role": "Veräußerung",
    # },
    # "Jean Keller (Castans Panoptikum)": {
    # "name": "Jean Keller (Castans Panoptikum)",
    # "date": "1887",
    # "role": "Vorbesitzer*in",
    # },
    # "José António de Oliveira (António Ole) (2013), Objektkünstler*in": {
    # "name": "José António de Oliveira (António Ole)",
    # "date": "2013",
    # "role": "Objektkünstler*in",
    # },
    # "J. Condt (?) (1875 (um)), Vorbesitzer*in": {
    # "name": "Serdu (?)",
    # "date": "1875 (um)",
    # "role": None,
    # },
    # "Laurance Austine Waddell (1905), Sammler*in', 'Gerson Simon (1905), Mäzen*atin'": {
    # "name": "Laurance Austine Waddell",
    # "date": "1905",
    # "role": "Sammler*in",
    # },
    # "Kaiser) Ch'ien Lung (China": {
    # "name": "Kaiser) Ch'ien Lung (China",
    # "date": "1711",
    # "role": None,
    # },
    # "KaiserlicheI Deutsches Konsulat in Tsi-nanfu (1909), Veräußerung": {
    # "name": "KaiserlicheI Deutsches Konsulat in Tsi-nanfu",
    # "date": "1909",
    # "role": "Veräußerung",
    # },
    # "Kanô Motonobu (狩野元信)": {
    # "name": "Kanô Motonobu (狩野元信)",
    # "date": "1476 - 1559",
    # "role": "Maler*in",
    # },
    # "Kawahara Keiga (川原慶賀)": {
    # "name": "Kawahara Keiga (川原慶賀)",
    # "date": "1786",
    # "role": "Maler*in",
    # },
    # "Kolonialzentralverwaltung (Reichsministerium für Wiederaufbau) (1921), Veräußerung": {
    # "name": "Kolonialzentralverwaltung (Reichsministerium für Wiederaufbau)",
    # "date": "1921",
    # "role": "Veräußerung",
    # },
    # "Korimex (S. Roho), Veräußerung": {
    # "name": "Korimex (S. Roho)",
    # "date": "1983",
    # "role": "Veräußerung",
    # },
    # "Max Heppner (Fa. Ludwig Glenk) (1905), Veräußerung": {
    # "name": "Max Heppner (Fa. Ludwig Glenk)",
    # "date": "1905",
    # "role": "Veräußerung",
    # },
    # "Matsumoto Setsutarô 松本節太郎 (根戸工房 Nedo Werkstatt), Hersteller*in": {
    # "name": "Matsumoto Setsutarô 松本節太郎 (根戸工房 Nedo Werkstatt)",
    # "date": "2004",
    # "role": "Hersteller*in",
    # },
    # "Nome Mokua (?) (1907), Vorbesitzer*in": {
    # "name": "Nome Mokua (?)",
    # "date": "1907",
    # "role": "Vorbesitzer*in",
    # },
    # "Ostasiatika Kunsthandel Berlin (Wolfgang Bock), Veräußerung": {
    # "name": "Ostasiatika Kunsthandel Berlin (Wolfgang Bock)",
    # "date": "1979",
    # "role": "Veräußerung",
    # },
    # "Ph(il).. Engelhardt (1903), Sammler*in": {
    # "name": "Ph(il). Engelhardt",
    # "date": "1903",
    # "role": "Sammler*in",
    # },
    # "R. SchmI D (1923), Veräußerung": {
    # "name": "R. Schmid",
    # "date": "1923",
    # "role": "Veräußerung",
    # },
    # "Rautenstrauch-Joest-Museum (Städtisches Museum für Völkerkunde Köln) (1901), Veräußerung": {
    # "name": "Rautenstrauch-Joest-Museum (Städtisches Museum für Völkerkunde Köln)",
    # "date": "1901",
    # "role": "Veräußerung",
    # },
    # "Rong Bao Zhai 荣宝斋 (Werkstatt), Hersteller*in": {
    # "name": "Rong Bao Zhai 荣宝斋 (Werkstatt)",
    # "date": "1982/1983",
    # "role": "Hersteller*in",
    # },
    # "Sanwa (Miwa) Yoshito (Yoshihito) 三輪義人": {
    # "name": "Sanwa (Miwa) Yoshito (Yoshihito) 三輪義人",
    # "date": 2011,
    # "role": "Kunsthandwerker",
    # },
    # "Schantung-Bergbaugesellschaft (Shandong-Bergbaugesellschaft) (1899 - 1914), Veräußerung": {
    # "name": "Schantung-Bergbaugesellschaft (Shandong-Bergbaugesellschaft)",
    # "date": "1899 - 1914",
    # "role": "Veräußerung",
    # },
    # "Schubarth (Schubart) (1911), Sammler*in": {
    # "name": "Schubarth (Schubart)",
    # "date": "1911",
    # "role": "Veräußerung",
    # },
    # "Schubarth (Schubart) (1911), Veräußerung": {
    # "name": "Schubarth (Schubart)",
    # "date": "1911",
    # "role": "Veräußerung",
    # },
    # "Serdu (?) (1875 (um)), Veräußerung": {
    # "name": "Serdu (?)",
    # "date": "1875 (um)",
    # "role": "Veräußerung",
    # },
    # "Stansky (Stanski), Veräußerung": {
    # "name": "Stansky (Stanski)",
    # "date": 1902,
    # "role": "Veräußerung",
    # },
    # "UchI Da Yoshiko (1924), Veräußerung": {
    # "name": "Uchida Yoshiko",
    # "date": 1924,
    # "role": "Veräußerung",
    # },
    # "Unbekannt, Veräußerung": {
    # "name": "Unbekannt",
    # "date": None,
    # "role": "Veräußerung",
    # },
    # "Unbekannt, Sammler*in": {
    # "name": "Unbekannt",
    # "date": None,
    # "role": "Sammler*in",
    # },

    for idx, case in enumerate(cases):
        prefix, name, date, role = _quad_split(case)
        match idx:
            case 0:
                assert prefix is None
                assert name == "Claus Schilling"
                assert date == "5.7.1871 (?) - 1946"
                assert role == "Sammler*in"
            case 1:
                assert prefix is None
                assert name == "Joachim Pfeil"
                assert date == "30.12.1857 - 12.3.1924"
                assert role == "Sammler*in"
            case 2:
                assert prefix is None
                assert name == "Kaiserliches Auswärtiges Amt des Deutschen Reiches"
                assert date == "1875"
                assert role == "Veräußerung"
            case 3:
                assert prefix == "Bezug unklar"
                assert name == "Paul Grade"
                assert date == "† 05.04.1894*"
                assert role is None
            case 4:
                assert prefix is None
                assert name == "A. Palamidessi (?)"
                assert date == "1939"
                assert role == "Veräußerung"
            case 5:
                assert prefix is None
                assert (
                    name
                    == "Academia Sinica (Nationale Akademie der Wissenschaften, Taiwan)"
                )
                assert date == "1962"
                assert role == "Veräußerung"
            case 6:
                assert prefix is None
                assert name == "Alex(ander) Siebold"
                assert date == "1872"
                assert role == "Veräußerung"
            case 7:
                assert prefix is None
                assert name == 'Augusta Kell ("Gulla") Pfeffer'
                assert date == "1887 - 1967"
                assert role == "Veräußerung"
            case 8:
                assert prefix is None
                assert name == "Baboo Mukharji (Muckharjee)"
                assert date == "1892"
                assert role == "Veräußerung"
            case 9:
                assert prefix is None
                assert name == "British Museum (Dep. of Ethnography)"
                assert date == "1893"
                assert role == "Veräußerung"


def tast_each_person3() -> None:
    from openpyxl import Workbook, load_workbook, worksheet

    wb = load_workbook("../sdata/Abschrift_HK_Afrika_III_C_Final.xlsx", data_only=True)
    ws = wb["Sheet1"]  # sheet exists already
    conf = {"project_dir": Path("../sdata"), "person_cache": "person_cache.toml"}
    person_data = open_person_cache(conf)

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        beteiligte = row[3].value
        if beteiligte is not None:
            beteiligtL = beteiligte.split(";")
            beteiligtL = [pk.strip() for pk in beteiligtL]
            for beteiligt in beteiligtL:
                if not beteiligt.isspace():
                    # print(f"{idx}:{beteiligt}")
                    for name, role, date in _each_person(beteiligt):
                        try:
                            pkIdL = person_data[name][date]
                        except:
                            print(f"test_each_person3 ERROR {name} {date}")
                        if not pkIdL:  # list is empty
                            print(f"{idx}: {name}{pkIdL}|{role}|{date}")
    # for count, (name, role, date) in enumerate(_each_person(beteiligte), start=1):
    #    print(f"{count}: {name} {role} {date}")
