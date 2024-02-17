"""
attach2 - this will become a tool to identify and fix for missing attachments 
"""

from mpapi.client import MpApi
from mpapi.module import Module
from mpapi.search import Search
from mpapi.constants import get_credentials, NSMAP
from MpApi.Record import Record  # tested?
from MpApi.Utils.Ria import RIA
from MpApi.Utils.Xls import Xls
from openpyxl import Workbook  # load_workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
from lxml import etree


class Attacher2:
    def __init__(
        self, *, cache: bool = False, excel_fn: Path = Path("attach.xlsx"), act=False
    ) -> None:
        self.cache = cache
        print(f"Using cache? {self.cache}")
        user, pw, baseURL = get_credentials()
        print(f"Logging in as {user} {baseURL}")
        self.client = MpApi(user=user, baseURL=baseURL, pw=pw)
        self.client2 = RIA(user=user, baseURL=baseURL, pw=pw)
        self.limit = -1
        print(f"Using limit {self.limit}")
        print(f"Using excel '{excel_fn}'")
        self.xls = Xls(path=excel_fn, description=self.desc())
        self.xls.save()
        self.act = act
        print(f"act? {self.act}")

    def desc(self) -> dict:
        desc = {
            "mulId": {
                "label": "Asset ID",
                "desc": "aus RIA",
                "col": "A",  # 0
                "width": 10,
            },
            "filename": {
                "label": "Dateiname",
                "desc": "aus RIA",
                "col": "B",  # 0
                "width": 40,
            },
            "location": {
                "label": "Pfad",
                "desc": "aus RIA",
                "col": "C",  # 0
                "width": 80,
            },
            "matches": {
                "label": "passender Pfad",
                "desc": "von Disk",
                "col": "D",  # 0
                "width": 80,
            },
            "uploaded": {
                "label": "schon hochgeladen?",
                "desc": "",
                "col": "E",  # 0
                "width": 10,
            },
        }
        return desc

    def ria(self) -> None:
        self.xls.raise_if_file()

        print("Getting data from RIA from ...")
        xfilter = (
            "/m:application/m:modules/m:module/m:moduleItem[@hasAttachments='false']"
        )
        results_fn = "debug_response.xml"

        print(self.cache)
        if self.cache:
            print(f"   cached response '{results_fn}'")
            m = Module(file=results_fn)
        else:
            print("   fresh query")
            q = self._get_query()
            q.validate(mode="search")
            print("Query validates, about to start search...")
            q.toFile(path="debug_query.xml")
            m = self.client.search2(query=q)
            new = m.filter2(xpath=xfilter)
            new.toFile(path=results_fn)
            print(f"Response written to disk '{results_fn}'")
        self._write_xlsx(new)

    def scandir(self) -> None:
        start_dir = Path(r"M:\MuseumPlus\Produktiv\Multimedia\EM\PLM-Dubletten\AmArch")

        self.xls.raise_if_no_file()
        self.ws = self.xls.get_or_create_sheet(title="Missing Attachments")

        filenames_from_excel = self._get_filenames_from_excel()

        c = 3
        for p in start_dir.glob("**/*"):
            if p.is_file:
                name = p.name
                if name in filenames_from_excel:
                    print(f"{c}:'{name}' on disk AND in excel list")
                    self._add_file_excel(p)
                print(f"files+dirs {c}", end="\r", flush=True)
            if c % 100_000 == 0:
                self.xls.save()
            c += 1
        self.xls.save()

    def up(self) -> None:
        """
        Attach the located files to the Multimedia recods in RIA.

        Loop through Excel and look in column "matches" (column D) which should contain
        absolute paths to the attachments, possibly multiple ones. Take the first one and
        attach that file to the asset record specified in column A.
        """
        self.xls.raise_if_no_file()
        self.ws = self.xls.get_or_create_sheet(title="Missing Attachments")
        self.xls.raise_if_no_content(sheet=self.ws)

        rno = 3  # one-based
        for row in self.ws.iter_rows(min_row=3):  # iter_rows one-based
            matches = row[3].value
            if matches is not None:
                mulId = row[0].value  # row is zero-based
                print(f"*** {rno}:{mulId}")
                if mulId is None:
                    raise Exception("ERROR: mulId missing!")
                # print(f"***{mulId}")
                matchesL = matches.split("; ")
                for match in matchesL:
                    print(f"{mulId} {match}")
                    if row[4].value == "x":
                        print(f"WARNING: Already uploaded according to Excel row {rno}")
                    else:
                        self._upload_attachment(path=match, mulId=mulId)
                        row[4].value = "x"
                        self.xls.save()  # after every upload
            rno += 1

    #
    # private
    #

    def _add_file_excel(self, p: Path) -> None:
        """
        Expect a path. Find the name in the Excel's first column and write the
        absolute path in that row into the last column.
        """
        hits = self._find_name_in_excel(p)
        if hits:
            for hit in hits:
                cl = f"D{hit}"
                if self.ws[cl].value is None:
                    print(f"hit: '{p.name}' {cl}")
                    self.ws[cl].value = str(p.absolute())
                else:
                    print(f"hit: '{p.name}' {cl} add")
                    self.ws[cl].value += f"; {str(p.absolute())}"

    def _find_name_in_excel(self, p: Path) -> list:
        hits = list()
        rno = 3  # search in col=2
        for row in self.ws.iter_cols(min_row=3, min_col=2, max_col=2):
            for cell in row:
                if cell.value == p.name:
                    hits.append(rno)
                rno += 1
        return hits

    def _get_filenames_from_excel(self) -> list:
        filenames_from_excel = list()
        rno = 3
        for row in self.ws.iter_cols(min_row=3, min_col=2, max_col=2):
            for cell in row:
                if cell.value is None:
                    continue
                fn = cell.value
                if fn not in filenames_from_excel:
                    filenames_from_excel.append(fn)
            rno += 1
        print(f"Filenames in Excel {filenames_from_excel}")
        return filenames_from_excel

    def _get_query(self) -> Search:
        q = Search(module="Multimedia")
        q.AND()
        q.addCriterion(
            operator="equalsField", field="__orgUnit", value="EMAmArchaologie"
        )
        q.addCriterion(operator="contains", field="MulOriginalFileTxt", value="pdf")
        return q

    def _upload_attachment(self, path: str, mulId: int) -> None:
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"ERROR: File {file} not found!")
        mulId = int(mulId)

        if not self.act:
            return

        m = self.client.getItem2(mtype="Multimedia", ID=mulId)
        if m:
            # print(f"asset ID {ID} exists in RIA")
            m.toFile(path=f"multimedia{mulId}.xml")  # debug not necessary
            ret = self.client2.upload_attachment(file=path, ID=mulId)
            print(f"return value after uploading attachment: {ret}")
            r = Record(m)
            # r.set_filename(path=file)
            # r.set_dateexif(path=file)
            r.set_size(path=path)
            m = r.toModule()
            r = self.client.updateItem4(data=m)
            print(f"updateItem4 multimedia-{mulId} return: {r}")
        else:
            raise Exception(f"multimedia ID not found online {mulid}")

    def _write_xlsx(self, data: Module) -> None:
        wb = self.xls.get_or_create_wb()
        self.ws = self.xls.get_or_create_sheet(title="Missing Attachments")
        self.xls.write_header(sheet=self.ws)
        rno = 3
        for moduleItemN in data.xpath("/m:application/m:modules/m:module/m:moduleItem"):
            mulId = moduleItemN.xpath("@id")[0]
            filename = moduleItemN.xpath(
                "m:dataField[@name='MulOriginalFileTxt']/m:value/text()",
                namespaces=NSMAP,
            )[0]
            try:
                location = moduleItemN.xpath(
                    "m:dataField[@name='MulOriginalFileLocationClb']/m:value/text()",
                    namespaces=NSMAP,
                )[0]
            except:
                location = None
            # print(f"{mulId} {filename} {location}")
            self.ws[f"A{rno}"] = mulId
            self.ws[f"B{rno}"] = filename
            if location is not None:
                self.ws[f"C{rno}"] = location
            rno += 1
        self.xls.save()


if __name__ == "__main__":
    pass
