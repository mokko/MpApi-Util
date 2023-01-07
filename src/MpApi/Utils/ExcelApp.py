"""
Let's make a base class that is inherited by every app that uses Excel, so we share some 
behavior.

So different apps that use differently structured Excel sheets can and have to them use 
in similar ways

from pathlib import path
class your_class(ExcelApp):

    self.excel_fn = Path("path/to.xlsx")
    self.wb = self.init_excel(path=self.excel_fn)
    self.save_excel(path=self.excel_fn) # relies on self.wb

So far this is near pointless, but perhaps I can still find a way to re-use significant 
parts of this class.
"""

from pathlib import Path
from openpyxl import Workbook, load_workbook


class ExcelApp:
    def init_excel(self, *, path: Path) -> Workbook:
        """
        Given a file path for an excel file, return the respective workbook
        or make a new one
        """
        if path.exists():
            print(f"* Loading existing excel: '{data_fn}'")
            #  load excel file
            return load_workbook(path)
        else:
            print(f"* Starting new excel: '{data_fn}'")
            self.wb = Workbook()

    def save_excel(self, path: Path) -> None:
        """
        Saves Excel file to disk at path; required self.wb.
        """
        print(f"Saving {self.fn} ...")
        self.wb.save(filename=path)
