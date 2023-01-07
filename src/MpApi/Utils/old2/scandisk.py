import re
from pathlib import Path
import openpyxl
from typing import Optional


class Scandisk:
    def _extractIdentNr(self, *, path: Path) -> Optional[str]:
        """
        TODO: We will need multiple identNr parsers so we have to find a way to
        configure that. Probably a plugin architecure
        """
        stem = str(path).split(".")[0]
        # stem = path.stem # assuming there is only one suffix
        # print (stem)
        m = re.search(r"([\w ,.-]+)\w*-KK", stem)
        # print (m)
        if m:
            return m.group(1)

    def _init_sheet(self):
        ws = self.wb.active
        ws.title = "prepareAsset"

        ws["A1"] = "Dateiname"  # von Verzeichnis
        ws["B1"] = "Signatur"  # aus Dateiname
        ws["C1"] = "schon hochgeladen?"  # aus RIA
        ws["D1"] = "objId"  # aus RIA
        ws["E1"] = "ganzer Pfad"  # aus Verzeichnis

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 100

        return ws

    def _info_from_disk(self, *, Dir):
        """
        Make this work when excel already filled in

        When info is filled in already, die immediately.

        """
        print("* Scanning source dir: {Dir}")
        lc = 2  # start writing in 2nd line

        for path in Path(Dir).rglob("*-KK*"):
            print(path)
            # if not self.exists_in_excel(path=path):
            self._per_line(line=lc, path=path)
            lc += 1
        self._save_data()

    def _per_line(self, *, line: int, path: Path) -> None:
        # specific to this scandisk task of prepare command
        filename = path.name
        identNr = self._extractIdentNr(path=path)

        ws = self.ws
        ws[f"A{line}"] = filename
        ws[f"B{line}"] = identNr
        ws[f"E{line}"] = str(path)

    def _save_data(self):
        """
        Saves Excel file to disk at self.fn
        """
        print(f"Saving {self.fn} ...")
        self.wb.save(filename=self.fn)

    #
    #
    #

    def scan_disk(self, *, Dir):
        if self.ws.max_row > 1:
            raise Exception("Error: Scan dir info already filled in")
        self._info_from_disk(Dir=Dir)
