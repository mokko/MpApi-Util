"""
Mover - moves files that are already in RIA to storage location.

mover init	    initialize Excel
mover scandir   recursively scan a dir
mover move      do the actual moving of the files

"""

from datetime import datetime
from mpapi.constants import get_credentials
from MpApi.Utils.BaseApp import BaseApp
from MpApi.Utils.Ria import RIA
from MpApi.Utils.Xls import Xls, ConfigError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
import re
import shutil
from tqdm import tqdm

excel_fn = Path("mover.xlsx")
red = Font(color="FF0000")
# parser = etree.XMLParser(remove_blank_text=True)
teal = Font(color="008080")


class Mover(BaseApp):
    def __init__(self, *, limit: int = -1):
        """
        breaks the go loop after number of items
        limit counts rows in Excel file, so limit < 3 is meaningless
        """
        self.limit = self._init_limit(limit)
        user, pw, baseURL = get_credentials()
        self.client = RIA(baseURL=baseURL, user=user, pw=pw)

        self.xls = Xls(path=excel_fn, description=self.desc())
        self.wb = self.xls.get_or_create_wb()

    def desc(self) -> dict:
        desc = {
            "filename": {
                "label": "Dateiname",
                "desc": "aus Verzeichnis",
                "col": "A",
                "width": 20,
            },
            "fn_exists": {
                "label": "Assets mit diesem Dateinamen",
                "desc": "mulId(s) aus RIA",
                "col": "B",
                "width": 20,
            },
            "fn_exists_orgUnit": {
                "label": "Assets mit diesem Dateinamen (orgUnit)",
                "desc": "mulId(s) aus RIA",
                "col": "C",
                "width": 20,
            },
            "move": {
                "label": "Bewegen?",
                "desc": "zu Backup Verzeichnis",
                "col": "D",
                "width": 8,
            },
            "notes": {
                "label": "Notizen",
                "desc": "werden nicht automatisch überschrieben",
                "col": "E",
                "width": 12,
            },
            "moved": {
                "label": "Bewegt",
                "desc": "",
                "col": "F",
                "width": 8,
            },
            "relpath": {
                "label": "relativer Pfad",
                "desc": "aus Verzeichnis",
                "col": "G",
                "width": 30,
            },
            "fullpath": {
                "label": "absoluter Pfad",
                "desc": "aus Verzeichnis",
                "col": "H",
                "width": 40,
            },
            "targetpath": {
                "label": "Zielpfad",
                "desc": "",
                "col": "I",
                "width": 40,
            },
        }
        return desc

    # perhaps this method should live somewhere else with other operations on paths
    # suspicious_asset_name
    def is_suspicious(self, fn: str) -> bool:
        p = Path(fn)
        if fn is None:
            return True
        elif fn.startswith("DSC") or fn.startswith("IMG_"):
            # default camera filenames are not sufficiently unique
            return True
        elif re.match(r"\d+", p.stem):
            return True
        elif p.stem.isspace():
            return True
        elif re.match(r"\w+{1:3}", p.stem):
            return True
        return False

    def init(self):
        """
        Creates a pre-structured, but essentially empty Excel file for configuration
        and logging purposes.

        Don't overwrite existing Excel file.
        """

        self.xls.raise_if_file()
        ws = self.xls.get_or_create_sheet(title="Dateien")
        self.xls.write_header(sheet=ws)
        self._make_conf()
        self.xls.save()

    def move(self):
        self._check_move()
        for c, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            if c["move"].value == "x" and c["moved"].value is None:
                if c["targetpath"].value is None:
                    # self.xls.save()
                    # This case should no longer be frequent
                    self._warning(
                        f"F{rno}",
                        "WARNING: Move says move, but targetpath has no info!",
                    )
                    # raise SyntaxError(
                    #    "ERROR: Move says move, but targetpath has no info!"
                    # )
                fro = Path(c["relpath"].value)
                print(f"{rno}/{self.ws.max_row}: {fro}")
                if c["targetpath"].value is not None:
                    to = Path(c["targetpath"].value)
                    try:
                        self._move(fro, to, rno, c)
                    except KeyboardInterrupt:
                        self.xls.request_shutdown()
                    else:
                        self.xls.set_change()
                else:
                    print("WARNING: target path is None")
            if rno % 500 == 0 and self.xls.changed:  # save every so often
                self.xls.backup_if_change()
                self.xls.save_if_change()
            self.xls.shutdown_if_requested()
        self.xls.backup_if_change()
        self.xls.save_if_change()

    def scandir(self):
        """
        I dont want to fill in targetpath if move != x
        Now, I do want to fill in targetpath if move != x
        """
        # check if excel exists, has the expected shape and is writable
        self._check_scandir()
        print(f"   filemask: {self.filemask}")

        c = 3
        with tqdm(total=self.ws.max_row - 2, unit=" files") as pbar:
            for p in Path().glob(self.filemask):
                # print(f"S{p}")
                p_abs = p.absolute()
                p_abs_str = str(p_abs)
                if p.name.startswith(".") or p.name.startswith("~") or p == excel_fn:
                    continue
                elif p.suffix in (".lnk"):
                    continue
                elif p.is_dir():
                    continue
                elif p.name.lower() == "thumbs.db" or p.name.lower() == "desktop.ini":
                    continue
                if self.exclude_dirs is not None:
                    for each in self.exclude_dirs:
                        if p_abs_str.startswith(each):
                            continue
                if self.xls.path_exists(path=p_abs, cno=7, sheet=self.ws):
                    # print(f"ff {p_abs.name}")
                    try:
                        pbar.update()
                    except KeyboardInterrupt:
                        self.xls.request_shutdown()
                else:
                    # print("new path")
                    try:
                        self._scan_per_file(path=p, count=c)
                    except KeyboardInterrupt:
                        self.xls.request_shutdown()
                    if c % 1000 == 0:  # save every so often
                        self.xls.save()
                if self.limit == c:
                    print("* Limit reached")
                    break
                c += 1
                self.xls.shutdown_if_requested()
        self.xls.backup()
        self.xls.save()

    def wipe(self):
        self._check_move()
        self.xls.wipe(sheet=self.ws)
        self.backup()
        self.save()

    #
    # private
    #
    def _check_move(self) -> None:
        self.xls.raise_if_no_file()
        self.xls.save()

        self.ws = self.xls.get_or_create_sheet(title="Dateien")
        self.xls.raise_if_no_content(sheet=self.ws)

    def _check_scandir(self) -> None:
        self.xls.raise_if_no_file()
        self.xls.save()
        self.xls.backup()
        self.ws = self.xls.get_sheet(title="Dateien")
        # if we want to continue a scan that was interrupted e.g b/c it reached its limit
        # we can't have this check
        # self.xls.raise_if_content(sheet=self.ws)
        self.orgUnit = self.xls.get_conf(cell="B2")  # can be None

        conf_ws = self.wb["Conf"]
        if conf_ws["B1"].value is None:
            raise ConfigError("ERROR: Need target directory!")

        self.target_dir = Path(conf_ws["B1"].value)

        if conf_ws["B3"].value is None:
            self.filemask = "**/*"
        else:
            self.filemask = conf_ws["B3"].value

        if conf_ws["B4"].value is None:
            self.exclude_dirs = []
        else:
            exclude_str = conf_ws["B4"].value
            excludeL = exclude_str.split(";")
            self.exclude_dirs = [d.strip() for d in excludeL]

    def _make_conf(self) -> None:
        conf = {
            "A1": "target dir",
            "A2": "orgUnit",
            "C2": "z.B. EMArchiv, EMMusikethnologie, EMSudundSudostasien",
            "A3": "Filemask",
            "B3": "**/*.jpg",
            "C3": """vollständige Python filemask; rekursives Scannen kann dadurch ab- und angestellt werden.""",
            "A4": "Exclude Dirs",
            "B4": "Andere Dokumente",
            "C4": """Mehrere Verzeichnisse durch ; trennen. Angegebene Verzeichnisse werden ignoriert.""",
            "A5": "Erstellungsdatum",
            "B5": datetime.today().strftime("%Y-%m-%d"),
        }
        self.xls.make_conf(conf)

    def _move(self, fro: Path, to: Path, rno: int, c: dict) -> None:
        """
        Copy file at fro to the path at to, make directories at target and write success
        in Excel.
        """
        if fro.exists():
            # don't overwrite existing files
            # since files with same name can exist in muliple folders
            # it's quite possible that files with same name exist multiple times
            if to.exists():
                # should not happen, as conflicts should be resolved earlier
                self.ws[f"I{rno}"].font = red
                self._warning(f"F{rno}", f"WARNING: target location exists")
                # raise Exception(f"file exists already: '{to}'")
            else:
                if not to.parent.exists():
                    try:
                        to.parent.mkdir(parents=True)
                    except FileNotFoundError as e:
                        self._warning(f"F{rno}", f"FileNotFoundError {e}")
                        return None
                try:
                    shutil.move(fro, to)
                except PermissionError as e:
                    # self.ws[f"I{rno}"].font = red
                    self._warning(f"F{rno}", f"PermissionError {e}")
                except OSError as e:
                    # eg disk full
                    self.xls.save()
                    raise e
                except FileNotFoundError as e:
                    self._warning(f"F{rno}", f"FileNotFoundError {e}")
                else:
                    self.ws[f"I{rno}"].font = teal
                    c["moved"].value = "x"
        else:
            self._warning(f"F{rno}", f"does not exist (anymore)")

    def _scan_per_file(self, *, path: Path, count: int) -> None:
        """
        Writes to self.ws
        """
        c = self.xls._rno2dict(count, sheet=self.ws)
        # only write in empty fields
        self._write_filename(c, path)

        if c["relpath"].value is None:
            c["relpath"].value = str(path)
        if c["fullpath"].value is None:
            c["fullpath"].value = str(path.absolute())
        self._write_fn_exists(c, path)
        self._write_fn_exists_orgUnit(c, path)
        self._write_move(c)
        self._write_targetpath(c)

        print(f"{count}: {path.name} [{c['move'].value}] {path.parent}")

    def _warning(self, cell_label: str, msg: str) -> None:
        print(msg)
        self.ws[cell_label].value = msg
        self.ws[cell_label].font = red

    def _write_filename(self, c, path):
        if c["filename"].value is None:
            c["filename"].value = path.name

        if self.is_suspicious(path.name):
            c["filename"].font = red

    def _write_fn_exists(self, c, path):
        if c["fn_exists"].value is None:
            if self.is_suspicious(path.name):
                c["fn_exists_orgUnit"].value = "None"
                c["fn_exists_orgUnit"].font = red
                return
            idL = self.client.fn_to_mulId(fn=path.name, orgUnit=None)
            if len(idL) == 0:
                c["fn_exists"].value = "None"
            else:
                c["fn_exists"].value = "; ".join(idL)

    def _write_fn_exists_orgUnit(self, c, path):
        if self.orgUnit is not None:
            if c["fn_exists_orgUnit"].value is None:
                if self.is_suspicious(path.name):
                    c["fn_exists_orgUnit"].value = "None"
                    c["fn_exists_orgUnit"].font = red
                    return
                idL = self.client.fn_to_mulId(fn=path.name, orgUnit=self.orgUnit)
                if len(idL) == 0:
                    c["fn_exists_orgUnit"].value = "None"
                else:
                    c["fn_exists_orgUnit"].value = "; ".join(idL)
        if c["fn_exists"].value != "None" and c["fn_exists_orgUnit"].value == "None":
            c["fn_exists"].font = red

    def _write_move(self, c):
        def is_number(x):
            if x is None:
                return False
            try:
                float(x)
                return True
            except ValueError:
                return False

        def is_a_list(x):
            if x is None:
                return False
            if ";" in x:
                return True

        if c["move"].value is None:
            if self.orgUnit is None:
                reference = c["fn_exists"]
            else:
                reference = c["fn_exists_orgUnit"]

            if is_number(reference.value):
                c["move"].value = "x"
            elif is_a_list(reference.value):
                c["move"].value = "x"
            else:
                c["move"].value = None

    def _write_targetpath(self, c):
        """
        I used to only write targetpath for condidates; now we write targetpath
        if there is a filename.
        """
        if c["filename"].value is not None:
            fro = Path(c["relpath"].value)
            to = self.target_dir / fro
            while to.exists():
                to = self._plus_one(to)
            c["targetpath"].value = str(to)
            c["targetpath"].font = teal
