"""
NAME rename - Quick and dirty rename tool with Excel proofing 
    Rename some files on the disk according to a pattern using an Excel 
    doc as intermediary step for manual proof reading.

    The first instance of this script was supposed to work on pairs of
    tiff and jpg files. Do we really need those jpgs in RIA? Or do we work 
    solely on tifs and eliminate the jpgs?

CONFIGURATION
    Where to we place the configuration information? In a plugin? Perhaps we 
    start with a quick and dirty version where we write the configuration in the 
    program file and fix that later if we ever need this script again.
    
    Maybe then we'll put it in $HOME/.bkp

CLI INTERFACE
    rename -s src -d dest -x o.xslx	: scan the specified directory and write the proposed 
                              changes to o.xsl
    rename -e -x o.xslx		: read o.xslx and execute the changes specified in the Excel
    
 EXCEL Format
    column A: old name (without path) 
    column B: new name
    column C: absolute orig path
    column D: new full path

    Should we switch to relative paths to be less redundant?

 FILE CACHE IN DICT
    Do we use a file cache to speed up the process? Only if we need jpg sisters
    
    absolute path is unique, so has to be key
    
    abspath: basename
         
"""
from openpyxl import load_workbook, Workbook
from pathlib import Path
from shutil import copyfile
from typing import Optional

# QUICK AND DIRTY CONF
src_dir = "\\pk.de\smb\Mediadaten\Projekte\AKU\MDVOS\Konvolute\EM_Ozeanien_Deterts_Dias_Hauser-Schäublin\02_Medien\TIFF\VIII Oz K 214\F1_01_TIFF"
start_no = 666  # -> VIII B 666.tif
dest_dir = "//pk.de/smb/Mediadaten/Projekte/AKU/MDVOS/data"


class Rename:
    def __init__(self) -> None:
        self.cache = {}
        self.lowest_no = start_no

    def execute(self, *, xls_fn):
        self.wb = load_workbook(filename=xls_fn)
        self.ws = self.wb.active
        line_count = 2
        for rno in range(2, self.ws.max_row):
            src = self.ws.cell(row=rno, column=3)
            dst = self.ws.cell(row=rno, column=4)
            print(f"{src} -> {dst}")
            # copyfile (src, dest)

    def scan(self, *, src_dir: str, dest_dir: str, xls_fn: str) -> None:
        """
        Scan specified directory for VIII B*.tif recursively, write results
        to Excel file.

        I think we should ignore the *.jpgs. There is a danger that we make mistakes
        and it multiplies the work. Let's go for the easiest version instead.
        If we decide we need jpgs, we need to rename them at the same time,
        Let's call them jpg sisters.
        """

        srcP = Path(start_dir)
        if not srcP.is_dir():
            raise SyntaxError("Source does not exist or is not a directory!")

        dstP = Path(dest_dir).resolve()
        if not dstP.is_dir():
            raise SyntaxError("Destination does not exist or is not a directory!")

        print(f"*About to scan {srcP}; destination is {dstP}")

        self._initXls()

        # for mask in r'VIII B*.tif': # , r'VIII B*.jpg'
        cnt = 2
        for p in Path(srcP).rglob(r"VIII B*.tif"):
            self._2xls(count=cnt, path=p, dst=dstP)
            # for each tif we may want to add the jpg sister
            cnt += 1
        wb.save(filename=xls_fn)

    #
    # privates
    #

    def _2xls(self, *, count, path, dstP):
        """
        For a given file path, fill out one row in the Excel file
        """
        absp = path.resolve()
        new_fn = f"VIII B {self.lowest}.tif"
        new_abs = dstP.joinpath(new_fn)
        self.lowest += 1
        print(f"{absp} -> {new_fn}")
        self.ws[f"A{count}"] = str(absp.name)  # orig name
        self.ws[f"B{count}"] = new_fn  # new name
        self.ws[f"C{count}"] = str(absp)  # orig abs path (unique)
        self.ws[f"D{count}"] = new_abs  # new absolute path

    def _initXls(self) -> None:
        """
        Just init xls file and first row with headinga
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "bpk rename"

        self.ws["A1"] = "Orig name"
        self.ws["B1"] = "New name"
        self.ws["C1"] = "Orig (absolute)"
        self.ws["D1"] = "New (absolute)"

    def _lookup(self, *, needle: str) -> Optional[str]:
        """
        Search the self.cache for a needle and return absolute path if needle
        exists or None if it doesn't.

        NOT TESTED
        """

        for key in self.cache:
            value = self.cache[key]
            if value == needle:
                return key
        return None

    def _mk_cache(self, *, start_dir: str) -> None:
        """
        Scan a dir recursively and write results in memory cache

        abspath : name

        We assume that abspath is unique, but that name is not necessarily unique.

        CURRENTLY NOT USED
        """

        srcP = Path(start_dir)

        print(f"Making jpg cache for {start_dir}")

        # for mask in r'VIII B*.tif', r'VIII B*.jpg':
        for p in Path(srcP).rglob(r"VIII B*.jpg"):
            absp = p.resolve()
            self.cache[str(absp)] = str(absp.name)
        # print (self.cache)
