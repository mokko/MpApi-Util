"""
Should emulate the hotfolder eventually. That is we

(a) read in a configuration from an Excel file
(b) process an input directory (non recursively),
(c) create new multimedia (=asset) records from a template
(d) upload/attach files to a multimedia record
(e) create a reference (usually from object to multimedia record)
(f) potentially set Standardbild

We no longer move successfully uploaded files; instead we record of state in the
Excel file.
NEW: I changed my mind again. We do move successful uploads again to easily see
what problems remain. This time to static directory. I wonder if that should be
configurable.

EXCEL ::: RIA
1)	IdentNr
2)	Sachbegriff
3)	Beteiligte (Rolle)
4)	Erwerb. Datum
5)	Erwerb. Nr.
6)	Erwerbung von
7)	Geogr. Bezüge
8)	Obj. Referenz A
9)	Obj. Referenz B
10)	 Inventarnotiz: wird das Feld jemals durch uns gefüllt
"""

from datetime import datetime
from lxml import etree
from mpapi.constants import get_credentials
from mpapi.module import Module
from MpApi.Record import Record  # should be MpApi.Record.Multimedia
from MpApi.Utils.BaseApp import BaseApp, ConfigError
from MpApi.Utils.logic import (
    extractIdentNr,
    extract_weitereNr,
    identNrParserError,
    is_suspicious,
    whole_for_parts,
)  # has_parts,
from MpApi.Utils.Ria import RIA
from MpApi.Utils.Xls import Xls

from openpyxl.styles import Font
from pathlib import Path
from PIL import Image
from PIL.ExifTags import Base as ExifBase

import shutil
from typing import Optional
from tqdm import tqdm

# adding number of fields to prevent accidental overwriting of old versions
excel_fn = Path("upload15.xlsx")
bak_fn = Path("upload15.xlsx.bak")  # should go away
parser = etree.XMLParser(remove_blank_text=True)
red = Font(color="FF0000")
teal = Font(color="008080")
green = Font(color="00FF00")

IGNORE_NAMES = (
    "checksum.md5",
    "desktop.ini",
    "debug.xml",
    "prepare.xlsx",
    "prepare.log",
    "prepare.ini",
    str(excel_fn),
    str(bak_fn),
    "thumbs.db",
)
IGNORE_SUFFIXES = (".py", ".ini", ".lnk", ".tmp")


class AssetUploader(BaseApp):
    def __init__(self, *, limit: int = -1, offset: int = 3) -> None:
        self.limit = self._init_limit(limit)
        print(f"Using limit {self.limit}")
        self.offset = int(offset)  # set to 3 by default to start at 3 row
        user, pw, baseURL = get_credentials()
        self.client = RIA(baseURL=baseURL, user=user, pw=pw)
        self.objIds_cache: dict[str, str] = {}
        self.xls = Xls(path=excel_fn, description=self.desc())

    def desc(self) -> dict:
        desc = {
            "filename": {
                "label": "Asset Dateiname",
                "desc": "aus Verzeichnis",
                "col": "A",  # 0
                "width": 20,
            },
            "identNr": {
                "label": "IdentNr",
                "desc": "aus Dateinamen",
                "col": "B",  # 1
                "width": 15,
            },
            "wNr": {
                "label": "Weitere Nr",
                "desc": "aus Dateinamen",
                "col": "C",  # 1
                "width": 15,
            },
            "asset_fn_exists": {
                "label": "Assets mit diesem Dateinamen",
                "desc": "mulId(s) aus RIA",
                "col": "D",  # 2
                "width": 15,
            },
            "objIds": {
                "label": "objId(s) aus RIA",
                "desc": "exact match für diese IdentNr",
                "col": "E",  # 3
                "width": 15,
            },
            "parts_objIds": {
                "label": "Geschwister",
                "desc": "für diese IdentNr",
                "col": "F",  # 4
                "width": 20,
            },
            "whole_objIds": {
                "label": "Ganzes objId",
                "desc": "exact match für diese IdentNr",
                "col": "G",  # 5
                "width": 20,
            },
            "ref": {
                "label": "Objekte-Link",
                "desc": "automat. Vorschlag für Objekte-DS",
                "col": "H",  # 6
                "width": 9,
            },
            "notes": {
                "label": "Bemerkung",
                "desc": "für Notizen",
                "col": "I",  # 7
                "width": 20,
            },
            "photographer": {
                "label": "Fotograf*in",
                "desc": "aus Datei",
                "col": "J",  # 8
                "width": 20,
            },
            "creatorID": {
                "label": "ID Urheber*in",
                "desc": "aus RIA",
                "col": "K",  # 9
                "width": 20,
            },
            "fullpath": {
                "label": "absoluter Pfad",
                "desc": "aus Verzeichnis",
                "col": "L",  # 10
                "width": 90,
            },
            # "targetpath": {
            # "label": "nach Bewegen der Datei",
            # "desc": "wenn Upload erfolgreich",
            # "col": "L",  # 11
            # "width": 30,
            # },
            "attached": {
                "label": "Asset hochgeladen?",
                "desc": "wenn Upload erfolgreich",
                "col": "M",  # 12
                "width": 15,
            },
            "standardbild": {
                "label": "Standardbild",
                "desc": "Standardbild setzen, wenn noch keines existiert",
                "col": "N",  # 13
                "width": 5,
            },
        }
        return desc

    def go(self) -> None:
        """
        Do the actual upload based on the preparations in the Excel file

        (a) create new multimedia (=asset) records from a template
        (b) upload/attach files to an multimedia records
        (c) create a reference usually from object to multimedia record
        (d) update Excel to reflect changes
        (e) set Standardbild (if x in right place)

        Is it allowed to re-run go multiple time, e.g. to restart attachment? Yes!

        BTW: go is now called 'up' in command line interface.
        """

        self._check_go()  # raise on error

        # breaks at limit, but doesn't save on its own
        for cells, rno in self.xls.loop(
            sheet=self.ws, offset=self.offset, limit=self.limit
        ):
            # relative path; assume dir hasn't changed since scandir run
            print(f"{rno}: {cells['identNr'].value} up")
            if cells["ref"].value == "None":
                print(
                    "   object reference unknown, not creating assets nor attachments"
                )
                continue

            if cells["fullpath"].value is not None:
                p = Path(cells["fullpath"].value)
            else:
                print("Fullpath is missing")
                continue

            match cells["attached"].value:
                case "x":
                    print("File already uploaded")
                case "File not found":
                    print("File already marked as missing")
                case _:
                    self._go(cells=cells, rno=rno, p=p)
            self.xls.save_bak_shutdown(rno=rno, bak=10)
        # self.xls.save_if_change()

    def init(self) -> None:
        """
        Creates a pre-structured, but essentially empty Excel file for configuration
        and logging purposes.

        Don't overwrite existing Excel file.
        """

        self.xls.raise_if_file()
        self.xls.get_or_create_wb()
        ws = self.xls.get_or_create_sheet(title="Assets")
        self.xls.write_header(sheet=ws)
        self._make_conf()
        self.xls.save()

    def initial_offset(self) -> int:
        """
        Returns number of rows with x in int representing the first row in the Excel without x in field
        "attached" (aka "Asset hochgeladen").
        """
        self._init_wbws()

        # we need a loop that doesn't break on limit
        c = 3  # row counter
        for row in self.ws.iter_rows(min_row=3):  # start at 3rd row
            if row[12].value == "x":
                c += 1
        return c

    def photo(self):
        """
        Loop thru the excel entries and lookup ids for fotographers/creators if no ID
        filled in yet.
        """
        self._init_wbws()
        for cells, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            self._photo(cells)
        self.xls.save()

    def scandir(self, *, Dir: Optional[Path] = None, offset: int = 0) -> None:
        """
        Scans local directory and enters values for each file in the Excel

        It is possible to re-run scandir. While re-running files in list that no longer
        exist will be deleted from the list and new files on disk will be added at the
        end of the list. The upshot is that user can rename files on disk or delete
        rows in Excel to re-index files by a scandir re-run.

        add new files, manually delete rows from Excel and
        to update the table by re-running scandir.
        """
        self.filemask: str  # make mypy happy
        self._check_scandir()
        print(f"Scanning {self.filemask}")
        # fast-forward cache: jump over all the files that have
        # already been attached according to Excel
        attached_cache = self._attached_cache()
        print("   in case of substantial file changes, create new Excel sheet")
        self.xls.save()

        print("Preparing file list...")
        file_list = list()  # set not necessary because every file only one time
        chunk_size = self.limit - offset
        print(f"   chunk size: {chunk_size}")
        with tqdm(total=chunk_size + len(attached_cache), unit=" files") as pbar:
            for idx, p in enumerate(Path().glob(self.filemask), start=1):
                # idx = counting files, no offset for headlines
                if (
                    p.name.startswith(".")
                    or p.name.startswith("debug")
                    or p.name.lower() in IGNORE_NAMES
                ):
                    # print("   exluding reason 1")
                    continue
                elif p.is_dir():
                    # print("   exluding reason 2: dir")
                    continue
                # we dont need to ignore suffixes if we look for *.jpg etc.
                elif self.filemask == "*" and p.suffix in IGNORE_SUFFIXES:
                    continue
                if str(p.absolute()) not in attached_cache:
                    file_list.append(p)
                    pbar.update()
                if self.limit == idx:
                    print("* Limit reached")
                    break

        print(f"Scanning sorted file list... {len(file_list)}")
        for idx, p in enumerate(sorted(file_list), start=1):
            print(f"scandir: {p}")
            rno = self.xls.path_exists(path=p.name, cno=0, sheet=self.ws)
            # rno is the row number in Assets sheet
            # rno is None if file not in list
            rno = self._file_to_list(path=p, rno=rno)
            self.xls.save_bak_shutdown(rno=idx, save=500, bak=1_000)
            if self.limit == idx:
                print("* Limit reached")
                break
        self.xls.save()

    def set_standardbild(self) -> None:
        """
        Set all items in the list as standardbild in the Excel file
        """
        self._check_scandir()
        print("Setting Standardbild for all entries")
        for c, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            c["standardbild"].value = "x"
            # print(f"{rno} {c['standardbild'].value}")
        self.xls.save()
        # raise SyntaxError

    def standardbild(self) -> None:
        """
        Loop thru Excel and only set standardbild if requested
        """
        print("Only setting Standardbild")
        self._check_scandir()
        for cell, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            # relative path; assume dir hasn't changed since scandir run
            # fn = cell["filename"].value

            print(f"{rno}: {cell['identNr'].value}")
            if cell["ref"].value is None:
                print("   no object reference cannot set standardbild")
                continue
            self._set_Standardbild(cell)
            self.xls.save_bak_shutdown(rno=rno, save=5, bak=10)

    def wipe(self) -> None:
        """
        Loop thru excel and delete all data rows. This re-creates the state of
        'upload init'. Can be useful for running another scandir.

        Beware of --limit.
        """
        self._init_wbws()
        self.xls.wipe(sheet=self.ws)

    #
    # private
    #
    def _attach_asset(self, *, path: str | Path, mulId: int) -> bool:
        """
        Attach an asset file (at path).
        """

        # if the file doesn't exist (anymore) it indicates major issues!
        if not Path(path).exists():
            raise FileNotFoundError("ERROR: Path doesn't exist. Already uploaded?")

        print(f"   attaching {path}")
        ret = self.client.upload_attachment(file=path, ID=mulId)
        # print(f"   success on upload? {ret}")
        if ret.status_code == 204:
            return True
        else:
            # should this raise an error?
            print("   ATTACHING FAILED (HTTP REQUEST)!")
            return False

    def _attached_cache(self) -> set:
        rno = 3  # row counter;
        cache = set()
        # loop without limit
        for row in self.ws.iter_rows(min_row=rno):  # start at 3rd row
            cells = self.xls._rno2dict(rno, sheet=self.ws)
            fullpath = cells["fullpath"].value
            attached = cells["attached"].value
            # print(f"{rno} {cells}")
            if attached == "x":
                cache.add(fullpath)
                # print(f"attached cache: {fullpath} {attached}")
            rno += 1
        print(f"Skipping files already uploaded ({len(cache)} files)")
        return cache

    def _check_go(self) -> None:
        """
        Checks requirements for go command. Raises on error.

        Saves workbook to self.wb
        """
        self._init_wbws()

        required = {
            "B1": "No templateID provided!",
            "B3": "orgUnit not filled in!",
            # B4 is now optional
        }
        self.xls.raise_if_conf_value_missing(required)
        self.move = self.xls.get_conf_true(cell="B4")

    def _check_scandir(self) -> None:
        """
        A couple of checks for scandir. Saves worksheet to self.wb.
        """
        # check if excel exists, has the expected shape and is writable
        self._init_wbws()
        self.xls.raise_if_not_initialized(sheet=self.ws)
        required = {
            "B5": "No filemask!",
            "B8": "No identNr parser provided!",
        }
        self.xls.raise_if_conf_value_missing(required)

        self.orgUnit = self.xls.get_conf(cell="B3")  # can be None
        self.filemask = self.xls.get_conf_required(cell="B5", default="*")
        self.ignore_suspicious = self.xls.get_conf_true(cell="B7")
        self.parser = self.xls.get_conf_required(cell="B8")
        if self.parser == "":
            raise ConfigError("Need identNr parser!")

        if self.parser == "iitm":
            print("loading prepare.xlsx for wNr to identNr mapping")
            self.ident_cache = {}
            from MpApi.Utils.prepareUpload import PrepareUpload

            pu = PrepareUpload()
            ident_xls = Xls(path="prepare.xlsx", description=pu.desc())
            asheet = ident_xls.get_sheet(title="Prepare")
            for cell, rno in ident_xls.loop(sheet=asheet):
                wNr = cell["wNr"].value
                identNr = cell["identNr"].value
                # print(f"{wNr} {identNr}")
                if wNr not in self.ident_cache:
                    self.ident_cache[wNr] = identNr

    def _create_from_template(
        self,
        *,
        fn: str | Path,
        objId: int,
        templateM: Module,
        creatorID: Optional[int] = None,
    ) -> Optional[int]:
        """
        Creates a new asset (Mulimedia) record in RIA by copying the template.
        Also fill in
        - object reference
        - filename
        - size

        CHANGES
        - Used to die if assetID was not defined; now just returns
        - Used to be called _make_new_asset
        """
        # print("enter _create_from_template")
        if objId is None or objId == "None":
            # Do we want to log this error/warning in Excel?
            raise SyntaxError(f"objId {objId} not allowed")
            print(f"moduleItemdId '{objId}' not allowed! Not creating new asset.")
            return None
        r = Record(templateM)
        r.add_reference(targetModule="Object", moduleItemId=objId)
        r.set_filename(path=fn)
        r.set_size(path=fn)
        if creatorID is not None:
            print(f"   creatorID {creatorID} _create_from_template")
            r.set_creator(ID=creatorID)
        newAssetM = r.toModule()
        # newAssetM.toFile(path="debug.template.xml")
        new_asset_id = self.client.create_asset_from_template(
            templateM=newAssetM,
        )
        return new_asset_id

    def _create_new_asset(self, cells: dict) -> None:
        """
        Copies a template specified in the configuration. Gets called during upload
        (go) phase.


        Writes ID of new asset in te field asset_fn_exists.
        """
        # print("_create_new_asset")
        if cells["asset_fn_exists"].value == "None":
            templateM = self._prepare_template()
            fn = cells["fullpath"].value
            if not Path(fn).exists():
                # fn had been found during a scandir process
                # if it is not found anymore at this time the file system has changed in
                # an important way which warrants an error and the user's attention.
                raise FileNotFoundError(f"File not found: '{fn}'")
            # print(f"fn: {fn}")
            creatorID = cells["creatorID"].value  # can be None
            # print(f"   template with creatorID {creatorID}")

            ref_objId = cells["ref"].value
            if ref_objId is None:
                print("WARNING: Ref objId not present! Not creating")
                return None

            new_asset_id = self._create_from_template(
                fn=fn,
                objId=ref_objId,
                templateM=templateM,
                creatorID=creatorID,
            )
            self.xls.set_change()
            cells["asset_fn_exists"].value = new_asset_id
            cells["asset_fn_exists"].font = teal
            print(f"   asset {new_asset_id} created")

    def _exif_creator(self, *, path: Path) -> Optional[str]:
        """
        Expect a pathlib path, try to read that file with exiv and return
        (a) a string with a single creator,
        (b) a semicolon separated list of creators as a str or
        (c) None if no creator could be found.

        A few file types are exempt from checking.
        """

        # known extensions that dont work with exif
        exclude_exts = (".jpg", ".exr", ".obj", ".pdf", ".xml", ".zip")
        if path.suffix.lower() in exclude_exts:
            print(f"\tExif: ignoring suffix {path.suffix}")
            return None

        try:
            with Image.open(str(path)) as img:
                img_data = img.getexif()
                # img_data = img.read_iptc() pyexiv2
        except:
            print("\tExif:Couldn't open for exif")
            return None

        try:
            return img_data[ExifBase.Artist.value]
            # old: img_data["Iptc.Application2.Byline"]
        except KeyError:
            print("\tExif:Didn't find photographer info")
            return None
        # else:
        #    return "; ".join(img_data[ExifBase.Artist.value])

    def _file_to_list(self, *, path: Path, rno=None):
        """
        If rno is None, add a new file to the end of the Excel list; else update the row
        specified by rno.

        This is for the scandir step.
        """
        # print(f"file_to_list '{path}'")
        if rno is None:
            rno = self.ws.max_row + 1  # max_row seems to be zero-based
        cells = self.xls._rno2dict(rno, sheet=self.ws)
        fullpath = path.absolute()  # .resolve() problems on UNC
        # only write in empty fields
        # relative path, but not if we use this recursively
        if cells["filename"].value is None:
            cells["filename"].value = path.name

        if cells["wNr"].value is None and self.parser == "iitm":
            cells["wNr"].value = extract_weitereNr(path)

        self._write_identNr(cells, path)

        if cells["fullpath"].value is None:
            cells["fullpath"].value = str(fullpath)
        # print (f"***{path}")
        self._write_asset_fn(cells, fullpath)

        # identNr_cell might be None, then we cant look up any objIds
        identNr = cells["identNr"].value
        if identNr is None:
            print(f"WARNING: identNr cell is empty!!!\n {path.name}")
            return

        if cells["objIds"].value is None:
            cells["objIds"].value = self._get_objIds(identNr=identNr)
        self._write_parts(cells)
        self._write_whole(cells)
        self._write_ref(cells)
        ref = cells["ref"].value
        print(f"   {rno}: {identNr} [{ref}]")
        self._write_photographer(cells, path)
        self._write_photoID(cells)
        return rno

    def _go(self, *, cells: dict, rno: int, p: Path):
        if p.exists():
            try:
                self._create_new_asset(cells)  # writes asset id to asset_fn_exists
                self._upload_file(cells, rno)
                self._set_Standardbild(cells)
            except KeyboardInterrupt:
                self.xls.request_shutdown()
        else:
            cells["attached"].value = "File not found"
            print(f"WARN: {p} doesn't exist (anymore)")

    def _init_wbws(self):
        self.xls.raise_if_no_file()
        # die if not writable so that user can close it before waste of time
        self.xls.save()
        try:
            self.wb
        except:
            self.wb = self.xls.get_or_create_wb()
        try:
            self.ws
        except:
            self.ws = self.wb["Assets"]

    def _get_mulId(self, *, fullpath: Path) -> int | str:
        """
        Expects a fullpath, returns mulId. Currently as str, should return int, I guess.
        """
        idL = self.client.fn_to_mulId(fn=str(fullpath.name), orgUnit=self.orgUnit)
        if len(idL) == 0:
            mulId = "None"
        else:
            mulId = "; ".join(idL)
        return mulId

    def _get_objIds(self, *, identNr: str) -> str:
        """
        For a given identNr:str return the objId or objIds. If multiple objIds match,
        they are returned as a string that is joined by "; ". If there is no match,
        the returned string is "None".

        The string is cached, so if you query the same identNr again the same run
        of the script, it doesn't need another TCP request.

        New
        - still in use for producing wholes (_write_wholes)
        - I tried strict=False, but this is too lose.
        """
        if identNr in self.objIds_cache:
            return self.objIds_cache[identNr]
        else:
            objIds = self.client.get_objIds(
                identNr=identNr, strict=True, orgUnit=self.orgUnit
            )
            self.objIds_cache[identNr] = objIds
            # print(f"   new objId from RIA [{objIds}]")
            return objIds

    def _make_conf(self) -> None:
        conf = {
            "A1": "templateID",
            "C1": "nur Asset ID , Hendryks default jpg 6697400",
            "A2": "verlinktes Modul",
            "B2": "Objekte",
            "C2": "noch nicht implementiert",
            "A3": "OrgUnit (optional)",
            "C3": """OrgUnits sind RIA-Bereiche in interner Schreibweise (ohne Leerzeichen). 
        Die Suche der existierenden Assets wird auf den angegebenen Bereich eingeschränkt. 
        Wenn kein Bereich angegenen, wird die Suche auch nicht eingeschränkt. Gültige 
        orgUnits sind z.B. EMArchiv, EMMusikethnologie, EMMedienarchiv, EMPhonogrammArchiv, EMSudseeAustralien""",
            "A4": "Move?",
            "B4": "False",
            "C4": """Wenn True werden hochgeladene Dateien in das Unterverzeichnis up bewegt; wenn False passiert nichts.""",
            "A5": "Filemask",
            "B5": "**/*.tif",
            "C5": "Filemask für rekursive Suche. Vorsicht *.tif ist nicht gleich *.tiff!",
            "A6": "Erstellungsdatum",
            "B6": datetime.today().strftime("%Y-%m-%d"),
            "A7": "Ignore suspicious?",
            "B7": "True",
            "C7": "Wenn der Wert True ist, werden verdächtige Dateinamen ignoriert und nicht weiter untersucht.",
            "A8": "IdentParser",
            "B8": "EM",
            "C8": "Algorithmus um Ident.Nr aus Dateiname zu extrahieren.",
        }
        self.xls.make_conf(conf)

    def _prepare_template(self) -> Module:
        try:
            return self.templateM
        except:
            ws2 = self.wb["Conf"]
            templateID = int(ws2["B1"].value)
            print(f"   template from asset {templateID}")
            self.templateM: Module = self.client.get_template(
                ID=templateID, mtype="Multimedia"
            )
            if not self.templateM:
                raise ValueError("Template not available!")
            # template.toFile(path=f".template{templateID}.orig.xml")
            return self.templateM

    def _upload_file(self, cells, rno) -> None:
        # only upload if not already uploaded and if we have an ID to upload to
        if (
            cells["asset_fn_exists"].value is None
            or cells["asset_fn_exists"].value == "None"
        ):
            print("   WARNING: no asset to attach to!")
            return None

        if cells["attached"].value is None:
            fn = cells["fullpath"].value

            ID = int(cells["asset_fn_exists"].value)
            if self._attach_asset(path=fn, mulId=ID):
                self.xls.set_change()
                cells["attached"].value = "x"
                self._move(fn)
        else:
            print("   asset already attached")

    def _move(self, fn: str) -> None:
        """
        Move file at location fn to subdir up
        """
        if self.move:
            src = Path(fn)
            new_dir = src.parent / "up"
            new_dir.mkdir(exist_ok=True)
            dst = new_dir / src.name
            if not src.exists():
                raise FileNotFoundError("Source file does not exist (anymore)")
            if not dst.exists():
                shutil.move(src, dst)
            else:
                print("Warning: dst exists already")

    def _set_Standardbild(self, c) -> Optional[int]:
        """
        Set asset as standardbild for known object; only succeeds if object has no
        Standardbild yet.
        """
        stdbild = c["standardbild"].value  # just a shortcut
        if c["asset_fn_exists"].value is None or c["asset_fn_exists"].value == "None":
            print("WARNING: No asset to set standardbild to")
            return None
        if stdbild is not None:
            if stdbild == "done":
                print("   standardbild says 'done' already")
            else:
                if stdbild == "x" and c["attached"].value == "x":
                    objId = int(c["ref"].value)
                    assetID = c["asset_fn_exists"].value
                    try:
                        mulId = int(assetID)
                    except:
                        # if multiple mulIDs, take the first
                        mulId = int(assetID.split(";")[0])
                    print("   setting standardbild")
                    r = self.client.mk_asset_standardbild2(objId=objId, mulId=mulId)
                    if r is not None and r.status_code == 204:
                        self.xls.set_change()
                        # print(f"xxx {r.status_code}")
                        # print("   setting column N to done")
                        c["standardbild"].value = "done"
                    else:
                        print("   NOT setting column N to done")
                    return 1
        return 0

    def _write_asset_fn(self, cells, fullpath):
        if cells["asset_fn_exists"].value is None:
            # if cache the known paths drastically reduces http requests
            cells["asset_fn_exists"].value = self._get_mulId(fullpath=fullpath)
            if cells["asset_fn_exists"].value != "None":
                print("   asset exists in RIA already")
                cells["attached"].value = "x"
                cells["attached"].font = red
                # red signifies that asset has already been uploaded, but it has not been
                # tested if asset is linked to any or correct object.
                # We need the x here to fast-forward during continous mode

    def _write_identNr(self, cells: dict, path: Path) -> None:
        wNr = cells["wNr"].value
        if cells["identNr"].value is None:
            if wNr is not None and self.parser == "iitm":
                try:
                    identNr = self.ident_cache[wNr]
                except:
                    raise SyntaxError(f"wNr not found: {wNr}")
            else:
                try:
                    identNr = extractIdentNr(
                        path=path, parser=self.parser
                    )  # returns Python's None on failure
                except identNrParserError:
                    cells["identNr"].font = red
                    identNr = ""

            if self.ignore_suspicious and is_suspicious(identNr=identNr):
                cells["identNr"].font = red
                return
            # currently only accepting identNrs that dont look suspicious
            # print(f"***{identNr=}")
            cells["identNr"].value = identNr

    def _write_parts(self, cells) -> None:
        """
        This is now used for siblings, not parts anymore. "Returns" values by writing
        to cells["parts_objIds"].

        We recently switched to to get_objIds_startswith internally. Now we have the
        problem that I B 100 a-k also finds I B 1009, so false siblings.

        we want to use the new get_objIds_startswith

        should return dict [identNr]: 12345; 1234

        dict = {
            "VII c 1234": [1234, 12345, 123456]
        }

        but it doesn't work yet
        """

        if cells["parts_objIds"].value is None and cells["objIds"].value == "None":
            # print(" _write_parts")

            identNr = cells["identNr"].value
            ident_whole = whole_for_parts(identNr)
            # print(f"+++{identNr}")

            # new version that adds a space after identNr
            IDs = self.client.get_objIds_startswith(
                orgUnit=self.orgUnit,
                identNr=ident_whole,
            )
            IDs2 = {}
            for objId in IDs:
                identNr = IDs[objId]
                if identNr.startswith(f"{ident_whole} "):
                    IDs2[objId] = identNr

            # format as string
            parts_str = ""
            for idx, objId in enumerate(IDs2, start=1):
                identNr = IDs2[objId]
                parts_str += f"{identNr}: {objId}"
                if idx < len(IDs2):
                    parts_str += "; "

            cells["parts_objIds"].value = parts_str
            # print(f" _write_parts: {parts_str}")
            if len(IDs2) == 0:
                cells["parts_objIds"].value = "None"

    def _write_whole(self, cells):
        # print("* write wholes")

        if cells["whole_objIds"].value is None:
            identNr = cells["identNr"].value
            ident_whole = whole_for_parts(identNr)
            # print(f"\t_write_whole {ident_whole}")
            objIds = self._get_objIds(identNr=ident_whole)  # new: strict=False
            if identNr != ident_whole:
                cells["whole_objIds"].value = f"{ident_whole}: {objIds}"
            else:
                cells["whole_objIds"].value = "None"

    def _write_photoID(self, cells):
        # print("\t_write_photoID")
        cname = cells["photographer"].value
        if cells["creatorID"].value is None and cname != "None" and cname is not None:
            print(f"\tlooking up creatorID '{cname}'")
            idL = self.client.get_photographerID(name=cname)
            # can be None, not "None". Since i may want to run 'upload foto' again after i have
            # added photographer to RIA's person module.
            if idL is None:
                print("\tcreatorID not found")
            else:
                idL_str = [str(ID) for ID in idL]
                cells["creatorID"].value = "; ".join(idL_str)
                # print(cells["creatorID"].value)

    def _write_photographer(self, cells, path):
        # if file already attached, we dont need to look for photographer again
        # assuming attached is either None or x, but not "" or anything
        if cells["photographer"].value is None and cells["attached"].value is None:
            # print("in photographer")
            creator = self._exif_creator(path=path)
            if creator is None:
                cells["photographer"].value = "None"
            else:
                cells["photographer"].value = creator

    def _write_ref(self, cells):
        """
        if asset_fn exists we assume that asset has already been uploaded
        We take the objId for a whole, and the first part objId if any.
        """
        if cells["ref"].value is None and cells["asset_fn_exists"].value == "None":
            objIds = cells["objIds"].value
            whole = cells["whole_objIds"].value
            siblings = cells["parts_objIds"].value

            # print(f"+++{siblings=}")
            if objIds != "None" and ";" not in str(objIds):
                # print("   taking ref from objIds...")
                cells["ref"].value = int(objIds)
                cells["ref"].font = teal
            elif whole != "None":
                wholeID = whole.split(":")[1].strip()
                # print(f"{wholeID=}")
                try:
                    cells["ref"].value = int(wholeID)
                except:
                    pass
                else:
                    cells["ref"].font = green
            if siblings != "None" and siblings is not None and siblings != "":
                print(f"***{siblings=}")
                if ";" not in siblings:
                    objId = int(siblings.split(": ")[1])
                    # print(f"NEW Ref: {objId}")
                    cells["ref"].value = objId
            # print(f" _write_ref: {cells['ref'].value} <-- {siblings}")
