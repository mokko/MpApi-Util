"""
Prepare for assets for upload (e.g with regular Bildimport)

At heart, this tool creates Object records for properly named asset files. New records 
are copied from a template record and get the identNr from the file.

With this tool we 
- recursively scan a directory
- filter for specific files (e.g. with "-KK" or only *.jpg)
- extract the identNr from filename
- check if asset has already been uploaded (sort of a "Dublette")
- lookup objId by identNr
- mark cases where extracted identNr are suspicious
- figure out cases where object record doesn't exit yet
- write results into spreadsheet
- for those cases create object records by copying a template record
- write the new identNr in the new record

CONCEPT
This tool is meant to be used by a person that we'll call the editor. The editor runs 
the script multiple times in different phases. The script writes its output into an
Excel file which we also use for configuration values. For each phase, the edior checks 
the results typically in the Excel file and, if necessary, corrects something. 

Prepare works on current working directory (aka pwd) and can work recursively, if so
configured (using the filemask in Conf[iguration] sheet of the Excel file).

There are now four phases
(0) init
(1) scandir 
(2) checkria and
(3) createobjects
(4) movedupes (optional)

   $ prepare scandir 
   $ prepare checkria  
   $ prepare createobjects    

After running scandisk
- check IdentNr have successfully been extracted
- if files are not named correctly/consistently, rename them
- check that schema ids have been identified; if necessary update schema db
- check if already uploaded results are plausible (current check is not exact)
- if there are a number of assets that have already been uploaded, consider moving them
  away using other util
- check cases where one file has multiple mulIds objIds
- if necessary delete Excel and re-run scandisk phase

After running checkria
- check if candidates are plausible
- revise candidates manually if desired

After running createobjects
- preserve Excel file for documentation; contains ids of newly created records

Update 
November 2023
- new step init to make it more similar to upload
April 2023
- changes to cli frontend
  (1) no more separate config file ('prepare.ini'), instead we use prepare.xlsx 
  (2) We now scan current working directory 
  (3) We get RIA credentials from $HOME/.RIA file  
  (4) 'prepare scandir' (without -p for phase)
  (5) hardcoded "prepare.xlsx" file name
  (6) scandisk renamed to scandir with alias init to be more similar to mover and upload
  (7) currently filemask is not configurable so always falls back to default
- changes inside
  (8) the indivdual checks from checkria into a single loop
  (9) got rid of old code
  (10) use the filename identNr parser from logic
TODO
- scandir should be re-runnable
- we should be able to set Standardbild
- Make use of the log file
"""

import configparser  # todo replace with toml in future
import logging
from lxml import etree
from mpapi.module import Module
from mpapi.constants import get_credentials
from MpApi.Utils.BaseApp import BaseApp
from MpApi.Utils.identNr import IdentNrFactory
from MpApi.Utils.logic import extractIdentNr, not_suspicious
from MpApi.Utils.Ria import RIA
from MpApi.Utils.Xls import Xls, ConfigError, NoContentError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
import openpyxl.cell.cell
from pathlib import Path
import re
import shutil
from typing import Any, Optional

red = Font(color="FF0000")


class PrepareUpload(BaseApp):
    def __init__(
        self,
        *,
        limit: int = -1,
    ) -> None:
        user, pw, baseURL = get_credentials()
        self.client = RIA(baseURL=baseURL, user=user, pw=pw)
        print(f"Logged in as '{user}'")
        self.limit = self._init_limit(limit)
        print(f"Using limit {self.limit}")
        self.xls = Xls(path="prepare.xlsx", description=self.desc())
        if self.xls.file_exists():
            print(f"* {self.xls.path} exists already")
        else:
            print(f"* About to make new Excel '{self.xls.path}'")

        self.wb = self.xls.get_or_create_wb()
        self.ws = self.xls.get_or_create_sheet(title="Prepare")

    #
    # public
    #

    def checkria(self) -> None:
        """
        We loop thru the Excel table, test if
        (a) assets with the given filename exist already
        (b) figure out the objId for the identNr
        (c) based on this information, fill in the candidate cell
        """
        self.xls.raise_if_no_content(sheet=self.ws)
        for cells, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            if cells["assetUploaded"] is not None and cells["identNr"] is not None:
                self.mode = "ff"
            else:
                self.mode = ""
            self._asset_exists_already(cells)
            self._objId_for_ident(cells)
            self._fill_in_candidate(cells)
            self._checkria_messages(cells, rno)

            if rno is not None and rno % 100 == 0:
                # dont save when in fast forward mode
                # if not self.mode == "ff":
                # self._save_excel(path=self.excel_fn)
                self.xls.save_if_change()
        self.xls.save_if_change()  # _save_excel(path=self.excel_fn)

    def create_objects(self) -> None:
        """
        Loop thru excel objId column. Act for rows where candidates = "x" or "X".
        For those, create a new object record in RIA using template record mentioned in
        the configuration (templateID).

        Write the objId(s) of the newly created records in candidate column.
        """

        self._check_create_objects()
        temp_str = self.xls.get_conf_required(cell="B1")
        ttype, tid = temp_str.split()
        ttype = ttype.strip()  # do i need to strip?
        tid_int = int(tid.strip())
        print(f"***template: {ttype} {tid}")

        self.xls.raise_if_no_content(sheet=self.ws)

        templateM = self.client.get_template(ID=tid_int, mtype=ttype)
        # templateM.toFile(path="debug.template.xml")

        for c, rno in self.xls.loop(sheet=self.ws, limit=self.limit):
            print(f"{rno} of {self.ws.max_row}")  # , end="\r" flush=True
            if c["identNr"].value is None:
                # without a identNr we cant fill in a identNr in template
                # should not happen, that identNr is empty and cadinate = x
                # maybe log this case?
                return
            if c["candidate"].value is not None:
                cand_str = c["candidate"].value.strip().lower()
                if cand_str == "x":
                    objIds_str = self._create_object(
                        identNrs=c["identNr"].value, template=templateM
                    )
                    c["objIds"].value = objIds_str
                    c["candidate"].value = None
                    self.xls.set_change()
                    if rno is not None and rno % 5 == 0:
                        # save almost immediately since likely to die
                        self.xls.save_if_change()
        self.xls.save_if_change()

    def desc(self) -> dict:
        desc = {
            "filename": {
                "label": "Dateiname",
                "desc": "aus Verzeichnis",
                "col": "A",
                "width": 20,
            },
            "identNr": {
                "label": "IdentNr",
                "desc": "aus Dateinamen",
                "col": "B",
                "width": 15,
            },
            "assetUploaded": {
                "label": "Asset hochgeladen?",
                "desc": "mulId(s) aus RIA",
                "col": "C",
                "width": 15,
            },
            "objIds": {
                "label": "objId(s) aus RIA",
                "desc": "für diese IdentNr",
                "col": "D",
                "width": 15,
            },
            "partsObjIds": {
                "label": "Teile/Ganzes",
                "desc": "objId für Teile/Ganzes",
                "col": "E",
                "width": 20,
            },
            "candidate": {
                "label": "Kandidat",
                "desc": "neue Objekte erzeugen?",
                "col": "F",
                "width": 7,
            },
            "notes": {
                "label": "Bemerkung",
                "desc": "",
                "col": "G",
                "width": 20,
            },
            "fullpath": {
                "label": "Pfad",
                "desc": "aus Verzeichnis",
                "col": "H",
                "width": 115,
            },
            "schema": {
                "label": "Schema",
                "desc": "aus IdentNr",
                "col": "I",
            },
            "schemaId": {
                "label": "SchemaId",
                "desc": "aus IdentNr",
                "col": "J",
            },
            "duplicate": {
                "label": "Duplikat",
                "desc": "aus IdentNr",
                "col": "K",
            },
        }
        return desc

    def init(self, conf: dict | None = None) -> None:
        self.xls.raise_if_file()
        wb = self.xls.get_or_create_wb()
        ws = self.xls.get_or_create_sheet(title="Prepare")
        self.xls.write_header(sheet=ws)
        self._make_conf(conf)  # write conf sheet
        self.xls.save()

    def scan_disk(self) -> None:
        """
        Recursively scan a dir (src_dir). List the files in an Excel file trying
        to extract the proper identNr.

        Filenames with suspicious characters (e.g. '-' or ';') are flagged by coloring
        them red.
        """
        self._check_scandir()
        c = 3  # start writing in 3rd line
        print(f"FILEMASK {self.filemask}, starting scan...")
        ignore_names = (
            "thumbs.db",
            "desktop.ini",
            "prepare.ini",
            "prepare.log",
            "prepare.xlsx",
        )
        file_list = list()
        src_dir = Path()  # Path(self.conf["src_dir"])
        print(f"* Scanning source dir: {src_dir}")
        for path in src_dir.glob(self.filemask):
            if path.is_dir():
                continue
            if path.name.startswith("."):
                continue
            if path.name.lower().strip() in ignore_names:
                continue
            if self.xls.path_exists(path=path.absolute(), cno=7, sheet=self.ws):
                # if absolute path is already in Excel ignore it
                continue
            file_list.append(path)
            c += 1
            if c == self.limit:
                print("* Limit reached")
                break
        print(f"* {len(file_list)} new files found")
        known_idents: set[str] = set()  # mark duplicates

        c = self.xls.real_max_row(sheet=self.ws) + 1
        for path in sorted(file_list):
            self._scan_per_row(c=c, path=path, known_idents=known_idents)
            print(f"sd {c} of {len(file_list)}")
            if c % 500 == 0:
                self.xls.save()
            c += 1
        self.xls.save()

    #
    # PRIVATE
    #

    def _asset_exists_already(self, c) -> None:
        """
        Mainly fills in the "already uploaded?" cell in Excel (column C).

        Checks if an asset with that filename exists already in RIA. If so, it lists the
        corresponding mulId(s); if not None

        If config value mv_dupes exists, move asset files to the directory from the
        mv_dupes config value.

        Creates the dupes dir if it doesn't exist.

        New:
        - The check is now specific to an OrgUnit which is the internal name of a Bereich
        (e.g. EMSudseeAustralien).
        - The search is not exact. RIA ignores Sonderzeichen like _; i.e. if we search
          for an asset with  name x_x.jog and we learn that this one exists already
          according to this method then we dont know if the filename is really x_x.jpg
          or any number of variants such as x__x.jpg.

        If the Excel cell is empty, we still need to run the test. If it has one, multiple
        mulIds or "None" we don't need to run it again.
        """
        ws2 = self.wb["Conf"]
        orgUnit = self.xls.get_conf(cell="B2")  # can return None
        if c["assetUploaded"].value == None:
            self.xls.set_change()
            idL = self.client.fn_to_mulId(fn=c["filename"].value, orgUnit=orgUnit)
            if len(idL) == 0:
                c["assetUploaded"].value = "None"
            else:
                c["assetUploaded"].value = "; ".join(idL)

    def _checkria_messages(self, c, rno) -> None:
        print(
            f"cr {c['filename'].value} -> {c['identNr'].value} {c['candidate'].value}"
        )
        print(f"{rno} of {self.ws.max_row}", end="\r", flush=True)

    def _check_create_objects(self) -> None:
        self.xls.save()
        required = {"B1": "Template config missing!"}
        self.xls.raise_if_conf_value_missing(required)
        self.filemask = self.xls.get_conf_required(cell="B3")
        print(f"Using filemask {self.filemask}")

    def _check_scandir(self) -> None:
        self.xls.raise_if_not_initialized(sheet=self.ws)
        self.xls.save()
        self.xls.raise_if_conf_value_missing({"B3": "Filemask"})
        self.filemask = self.xls.get_conf_required(cell="B3")
        print(f"Using filemask {self.filemask}")
        self.parser = self.xls.get_conf_required(cell="B4")
        identNrF = IdentNrFactory()
        self.schemas = identNrF.get_schemas()

    def _create_object(self, *, identNrs: str, template) -> str:
        identL = identNrs.split(";")
        objIds = set()  # unique list of objIds from Excel
        for ident in identL:
            identNr = ident.strip()
            # print(f"***trying to create new object '{identNr}' from template")
            new_id = self.client.create_from_template(
                template=template, identNr=identNr
            )
            # logging.info(f"new record created: object {new_id} with {identNr} from template")
            objIds.add(new_id)
        objIds_str = "; ".join(str(objId) for objId in objIds)
        return objIds_str

    def _fill_in_candidate(self, c) -> None:
        if c["schemaId"].value is None or c["schemaId"].value == "None":
            c["candidate"].font = red
        if c["candidate"].value is None:
            if (
                c["assetUploaded"].value == "None"
                and c["objIds"].value == "None"
                and c["partsObjIds"].value == "None"
                and c["duplicate"].value != "Duplikat"
                and c["schemaId"].value != "None"
                and not_suspicious(c["filename"].value)
            ):
                c["candidate"].value = "x"

    def _get_objIds(self, *, identNr: str, strict: bool) -> str:
        orgUnit = self.xls.get_conf(cell="B2")  # can return None
        for single in identNr.split(";"):
            ident = single.strip()
            objIdL = self.client.identNr_exists(
                nr=ident, orgUnit=orgUnit, strict=strict
            )
            if not objIdL:
                return "None"
            return self._rm_garbage("; ".join(str(objId) for objId in objIdL))
        return "None"

    def _get_objIds2(self, *, identNr: str, strict: bool) -> str:
        """
        Superloaded version of get_objIds that only lets real parts through. Not very
        fast, but since RIA cant search for Sonderzeichen there is no way around it.

        We could move the logic that identifies parts to the RIA module though. But
        we have to move the garbage eliminator as well. Not today.
        """
        orgUnit = self.xls.get_conf(cell="B2")  # can return None
        for single in identNr.split(";"):
            identNr = single.strip()
            resL = self.client.identNr_exists2(
                nr=identNr, orgUnit=orgUnit, strict=strict
            )
            if not resL:
                return "None"
            real_parts = []
            for result in resL:
                objId = result[0]
                identNr_part = self._rm_garbage(result[1])
                if f"{identNr} " in identNr_part:
                    real_parts.append(f"{objId} [{identNr_part}]")
            # if we tested some results, but didnt find any real parts
            # we dont want to test them again
            if not real_parts:
                return "None"
            return "; ".join(real_parts)
        return "None"

    def _make_conf(self, conf: dict | None = None) -> None:
        """
        Makes a conf sheet with default and used-supplied values.
        """
        default = {
            "A1": "template ID",
            "C1": "Format: Object 1234567",
            "A2": "orgUnit",
            "C2": """Schränke die ID Suche auf eine orgUnit (Bereich) ein. Optional. z.B. EMSudseeAustralien""",
            "A3": "Filemask",
            "C3": """Steuere scandir Prozess mit einem Muster, z.B. '**/*' oder '*.jpg'.""",
            "A4": "IdentNr Parser",
            "B4": "EM",
            "C4": "Welcher Logarithmus zum Parsen von Dateinamen in identNrn soll verwendet werden? (EM, Std)",
        }

        if conf is not None:
            conf = default | conf  # mix defaults and user-defined conf
        else:
            conf = default  # use only defaults

        self.xls.make_conf(conf)

    def _objId_for_ident(self, c) -> None:
        """
        Writes in two cells: objIds and candidate

        Lookup objIds for IdentNr. Write the objId(s) to Excel. If none is found, write
        the string "None".

        Also writes x in candidate cell if uploaded and objId cell both have "None";
        write y if schemaId is missing.
        """

        # in rare cases identNr_cell might be None, then we cant look up anything
        if c["identNr"].value is None:
            return

        if c["objIds"].value == None:
            c["objIds"].value = self._get_objIds(
                identNr=c["identNr"].value, strict=True
            )
            self.xls.set_change()

        # used to check if c["objIds"].value == "None"
        if c["partsObjIds"].value is None:
            # print ("Looking for parts")
            objIdL = self._get_objIds_for_whole_or_parts(identNr=c["identNr"].value)
            self.xls.set_change()

            if objIdL:
                c["partsObjIds"].value = "; ".join(
                    [str(x) for x in objIdL]
                )  # to make mypy happy
            else:
                c["partsObjIds"].value = "None"
            c["partsObjIds"].alignment = Alignment(wrap_text=True)

    def _scan_per_row(self, *, c: int, path: Path, known_idents: set) -> None:
        """
        c: row count
        specific to this scandisk task of prepare command

        writes to self.ws
        """
        identNr = extractIdentNr(path=path, parser=self.parser)
        identNrF = IdentNrFactory()
        print(f"   {path.name} -> {identNr}")
        self.ws[f"A{c}"] = path.name
        self.ws[f"B{c}"] = str(identNr)
        self.ws[f"H{c}"] = str(path.absolute())
        self.xls.set_change()

        if identNr is not None:
            schema = identNrF._extract_schema(text=identNr)
        else:
            schema = "None"
        self.ws[f"I{c}"] = schema

        try:
            schemaId = self.schemas[schema]["schemaId"]
            self.ws[f"J{c}"] = schemaId
        except:
            self.ws[f"J{c}"] = "None"
            self.ws[f"J{c}"].font = red

        if self._suspicious_characters(identNr=identNr):
            for x in "ABCDEF":
                self.ws[f"{x}{c}"].font = red
            print(f"WARNING: identNr is suspicious - file correctly named? {identNr}")
        # If the original files are misnamed, perhaps best to correct them instead of
        # adapting the parser to errors.

        if identNr in known_idents:
            self.ws[f"B{c}"].font = red
            self.ws[f"K{c}"] = "Duplikat"
            # print(f"Duplikat {identNr}")
        known_idents.add(identNr)

    def _suspicious_characters(self, *, identNr: str | None) -> bool:
        # print (f"***suspicious? {identNr}")

        msg = "suspicious_characters:"

        if identNr is None:
            # print ("return bc None")
            return True
        elif "  " in identNr:
            logging.info(f"{msg} double space {identNr}")
            return True
        elif "." in identNr:
            # TODO seems that identNr with . are not mrked
            logging.info(f"{msg} unwanted symbol {identNr}")
            return True
        elif " " not in identNr:
            logging.info(f"{msg} missing space {identNr}")
            return True
        elif "-a" in identNr:
            logging.info(f"{msg} combination -a {identNr}")
            return True
        elif identNr.count(",") > 1:
            logging.info(f"{msg} number of commas {identNr}")
            return True

        # print (" -> not suspicious")
        return False


if __name__ == "__main__":
    pass
