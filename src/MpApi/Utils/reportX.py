"""
ReportX writes reports describing the files in the current directory.

The report is written in Excel (xlsx). It's basically a list of files with some information
(size, mtime etc.).

"""
import datetime
from MpApi.Utils.logic import extractIdentNr
import MpApi.Utils.BaseApp
from MpApi.Utils.Xls import Xls, BaseApp, ConfigError
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles import Alignment, Font
from pathlib import Path

# from Typing import Optional
fast = False


class ReportX(BaseApp):
    def __init__(self, limit=-1) -> None:
        self.limit = int(limit)

    def write_report(self, fn: str) -> None:
        """
        We assume that the report is empty at the beginning, loop thru all files and enter
        each one into Excel table.
        """
        self.xls = Xls(path=fn)
        ws = self._init_report(path=xlsx_fn)
        ws2 = self.wb.create_sheet("Conf")
        rno = 2
        print("beginning recursive scandir")
        # probably not better if we sort it first since we potentially have to wait to long.
        # TODO: But then perhaps we should include identNr sort
        for p in Path().rglob("*"):
            if p.is_dir():
                print(f" dir: {p}")
                continue
            elif p.name.lower() == "thumbs.db" or p.name.lower() == "desktop.ini":
                continue
            identNr = extractIdentNr(path=Path(p.name))
            print(f"   {rno}: {p} -> {identNr}")
            ws[f"A{rno}"].value = p.name
            if not fast:
                ws[f"B{rno}"].value = int(p.stat().st_size / 1024)
                ws[f"C{rno}"].value = p.stat().st_mtime
            ws[f"D{rno}"].value = identNr
            ws[f"E{rno}"].value = str(p)
            ws[f"F{rno}"].value = str(p.absolute())
            if (rno / 10000).is_integer():
                # save periodically
                self._save_excel(path=xlsx_fn)
            rno += 1
            if self.limit == rno:
                break
        ws2["A1"].value = "done"
        self._save_excel(path=xlsx_fn)

    #
    # private
    #

    def _init_report(self, *, path: Path) -> worksheet:
        """
        Creates a new report and saves it in self.wb. Also returns the new first
        worksheet.
        """
        self.xls.raise_if_file()
        self.wb = self.xls.get_or_create_wb()
        now = datetime.datetime.now().strftime("%Y-%m-%d")
        ws = self.xls.get_or_create_sheet(now)
        print(f"new sheet {now}")
        ws["A1"] = "Dateiname"
        ws["B1"] = "Größe (KB)"
        ws["C1"] = "mtime"
        ws["D1"] = "IdentNr?"
        ws["E1"] = "rel. Verzeichnis"
        ws["F1"] = "Absoluter Pfad"

        for each in "A1", "B1", "C1", "D1", "E1", "F1":
            ws[each].font = Font(bold=True)

        ws.column_dimensions["A"].width = 17
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 50

        return ws
