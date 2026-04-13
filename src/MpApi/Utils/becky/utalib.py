"""
Script to import records from Excel for Kamerun project

Fields that have to fill in
- IdentNr
- Sachbegriff
- Beteiligte + Rolle
- Erwerb.Datum
- Erwerbungsart
- Erwerb. Nr.
- Erwerbung von
- Geogr. Bezug
- Obj. Referenz A
fraglich
- Obj. Referenz B
- Inventarnotiz

New:
Workflow where we log errors and dont create record with missing info, but run through

"""

import argparse
from datetime import datetime
import logging
from lxml import etree  # t y p e : i g n o r e

from mpapi.constants import get_credentials
from mpapi.module import Module
from mpapi.search import Search

from MpApi.Utils.Ria import RIA, init_ria, record_exists2, record_exists3
from MpApi.Utils.becky.write_xml import create_xml
from MpApi.Utils.Xls import Xls
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string
from pathlib import Path
import re
from rich import print as rprint
import tomllib

#
# CONFIGURATION
#

no_records_created = 0
verbose = 0  # false for off, true for on.


def create_record(*, row: tuple, conf: dict, act: bool) -> None:
    # print(">> Create record")
    # missing_info = False NOT HERE
    # ident_col = conf["fields"]["identNr"]["identNr"]

    recordM, missing_info = create_xml(conf=conf, row=row)
    ident = get_ident(conf, row)  # for messages
    print(f"DDD: {ident}")

    # print(recordM)
    if missing_info:
        msg = f"Not creating record in RIA '{ident}' since missing info"
        logging.error(msg)
        print(f">> {msg}")
    elif act:
        # we used to count also would-be created records without act
        global no_records_created
        no_records_created += 1
        objId = conf["RIA"].create_item(item=recordM)
        msg = f"Created record {objId} in RIA '{ident}'"
        logging.error(msg)
        print(f">> {msg}")
    else:
        print(f">> Not creating record in RIA '{ident}' (since no act)")

    raise Exception("Stop here!")


def dd(msg: str) -> None:
    """Debugging print messages"""

    if verbose:
        print(msg)


def get_ident(conf: dict, row: list) -> str:
    """
    Assuming you defined a cluser identNr with the field identNr, this returns the identNr
    for the current row.
    """
    ident_col = column_index_from_string(conf["fields"]["identNr"]["identNr"]) - 1
    ident = row[ident_col].value  # from Excel as str
    # rprint(f"{ident_col=} {ident=}")
    if ident is None:
        logging.warning(f"IdentNr is None {idx}; not processing this line")
        return None
    return ident


def go_display_record(line_number: int, *, conf: dict, ws: worksheet):
    import sys
    from rich.console import Console

    console = Console()
    print(f">> Display record {line_number}")
    for cluster in conf["fields"]:
        console.print(f"[blue]{cluster}[reset]:")
        for field in conf["fields"][cluster]:
            value = conf["fields"][cluster][field]
            if is_excel_column(value):
                column = value
                coord = f"{column}{line_number}"
                if ws[coord].value is not None:
                    excel = ws[coord].value
                    console.print(
                        f"   [green]{field}[reset]: [yellow]{excel}[reset] [red]{column}[reset]"
                    )
                else:
                    console.print(
                        f"   [green]{field}[reset] [red]{column}[reset] empty cell"
                    )
            else:
                console.print(f'   [green]{field}[reset]: "{value}" constant')

    sys.exit(0)
    # raise Exception("Stop here")


def init_log(*, act: bool, conf: dict, conf_fn: str, limit: int, offset: int) -> None:
    """
    Create a simple logger at file becky20250510-0956.log
    """
    # should we only log if we actually do something with act=True?
    # to avoid plethora of log files?
    if act is True:
        now = datetime.now()
        datetime_str = now.strftime("%Y%m%d-%H%M%S")
        log_fn = f"becky{datetime_str}.log"
        logging.basicConfig(
            filename=log_fn,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
        )
        # - %(name)s is currently not necessary
        logger = logging.getLogger(__name__)
        logger.info(f"becky started with {act=}, {offset=} and {limit=}")
        logger.info(f"loading Excel file '{conf['excel_fn']}'")
    else:
        logger = logging.getLogger(__name__)
        logger.addHandler(logging.NullHandler())


def is_excel_column(s: str) -> bool:
    """
    Test if a string is an Excel column (e.g. AAA) or just a regular string.
    Columns must be 1–3 uppercase letters only; highest column XFD.
    """
    if not isinstance(s, str) or len(s) == 0 or len(s) > 3:
        return False
    if not s.isalpha() or not s.isupper():
        return False
    return len(s) < 3 or s <= "XFD"


def log_print_info(msg: str) -> None:
    """
    log and print info message simultaneously
    """
    logger = logging.getLogger(__name__)
    logger.info(msg)
    print(f"   {msg}")


def per_row(*, idx: int, row: Cell, conf: dict, act: bool) -> None:
    # rprint(conf["fields"])
    ident = get_ident(conf, row)  # should it die on no ident? Die early?

    # font_color = row[0].font.color
    # if font_color and font_color.rgb == "FFFF0000":  # includes the alpha channel
    global no_records_created
    print(f"***[{no_records_created}]{idx}: {ident}")
    # record_exists2 is Hendryk's algorithm that uses schemata and fortlaufende Nummer
    # if m := record_exists2(ident=ident, conf=conf):
    # record_exists3 omits Bereich and simply uses IdentNr and exact match.
    if m := record_exists3(ident=ident, conf=conf):
        # Wollen wir hier Fehler loggen um Nachzuvollziehen, wo die Infos aus Excel
        # nicht eingetragen wurden? Nein. Nur loggen, wenn etwas in RIA verändert wird
        print(f"INFO Record '{ident}' exists already")
    else:
        create_record(row=row, conf=conf, act=act)
        if m > 1:
            logging.warning(
                f"Multiple identNr: More than one IdentNr exists already with this number {ident}"
            )


def prepare_fields(conf: dict) -> None:
    """
    Rewrite the fields so we have less work later. I am not sure when to do this. At this point early in the game
    It's efficient because we only have to do this part once. But then we will have to do the next part later.
    If we only do it later, we can do it for indivdual cells and we only have to do it once.
    """
    conf["fields2"] = {}
    for cluster in conf["fields"]:
        print(f"c:{cluster}")
        conf["fields2"][cluster] = {}
        conf["fields2"][cluster]["_cb"] = create_callback(cluster)
        for field in conf["fields"][cluster]:
            print(f"    {field}")
            conf["fields2"][cluster][field] = {}
            if is_excel_column(conf["fields"][cluster][field]):
                conf["fields2"][cluster][field]["col"] = conf["fields"][cluster][field]
            else:
                conf["fields2"][cluster][field]["constant"] = conf["fields"][cluster][
                    field
                ]
    # rprint(conf["fields"])
    rprint(conf["fields2"])

    # raise Exception("Stop here!")


def create_callback(name: str) -> str:
    """
    We receive the name of field from the configuration toml file and return the
    name of the Python function we want to call.
    """
    (name2, no) = cluster_splitter(name)
    print(f"{name2=}{no=}")
    prefix = "set_"  # if no number, default to overwrite template values
    if no == 1:
        prefix = "set_"
    if no > 1:
        prefix = "add_"
    new_name = f"{prefix}{name2}"
    return new_name


def cluster_splitter(field: str) -> tuple[str, int]:
    """
    Split off trailing number and return both separately. If number doesn't exist, return field name as is and a 0.
    """
    match = re.match(r"(.*?)(\d+)$", field)
    if match:
        field2, no = match.groups()
        return field2, int(no)
    else:
        return field, 0


def prepare_template(conf: dict) -> Module:
    """
    Get the template record from RIA and rewrite it to approximate upload form.
    """
    print(f">> Getting template from RIA Object {conf['template_id']}")
    templateM = conf["RIA"].get_template(ID=conf["template_id"], mtype="Object")
    templateM._dropFieldsByName(element="systemField", name="__uuid")
    templateM._dropAttribs(xpath="//m:moduleItem", attrib="id")
    return templateM


def uta_main(
    *,
    conf_fn: str,
    act: bool = False,
    limit: int = -1,
    offset: int = 2,
    display_record: int | None,
) -> None:
    conf = _load_conf(conf_fn)  # sets project_dir
    print(f">> Setting project_dir '{conf['project_dir']}'")

    wb = load_workbook(conf["excel_fn"], read_only=True)
    ws = wb[conf["sheet_title"]]  # sheet exists already

    if display_record:
        go_display_record(display_record, conf=conf, ws=ws)

    conf["RIA"] = init_ria()  # mpApi.util.Ria's client
    init_log(act=act, conf=conf, conf_fn=conf_fn, limit=limit, offset=offset)
    conf["templateM"] = prepare_template(conf)

    prepare_fields(conf)

    for idx, row in enumerate(ws.iter_rows(min_row=conf["excel_row_offset"]), start=2):
        dd(f"{idx=} {offset=}")
        if idx < offset:
            continue
        per_row(idx=idx, row=row, conf=conf, act=act)
        if limit == idx:
            print(f">> Limit reached {limit}")
            break


#
# more private
#


def _load_conf(conf_fn: str) -> dict:
    print(f">> Reading configuration '{conf_fn}'")
    with open(Path(conf_fn), "rb") as toml_file:
        conf = tomllib.load(toml_file)
    conf["project_dir"] = Path(__file__).parents[4] / "sdata"  # project_dir
    return conf
