"""
Script to import records from Excel for Kamerun project

Fields that have to fill in
- IdentNr
- Sachbegriff
- Beteiligte + Rolle
- Erwerb.Datum
- Erwerbungsart
- Erwerb. Nr.
- Erwerbung von
- Geogr. Bezug
- Obj. Referenz A
fraglich
- Obj. Referenz B
- Inventarnotiz

Todo:
- Why dont we need to provide a namespace?
- Put the template record into a group ("Kamerun 2024"), so all created records are in a group
"""

import argparse
from copy import deepcopy
from datetime import datetime
import logging
from lxml import etree  # t y p e : i g n o r e

from mpapi.constants import get_credentials
from mpapi.module import Module
from mpapi.search import Search

from MpApi.Utils.Ria import RIA, init_ria, record_exists2
from MpApi.Utils.becky.set_fields_Object import (
    set_ident,
    set_ident_sort,
    set_sachbegriff,
    set_beteiligte,
    set_erwerbDatum,
    set_erwerbungsart,
    set_erwerbNr,
    set_erwerbVon,
    set_geogrBezug,
    set_invNotiz,
    set_objRefA,
)
from MpApi.Utils.Xls import Xls
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
from pathlib import Path
import tomllib

#
# CONFIGURATION
#

hits = 1  # global variable, 1-based


def becky_main(*, conf_fn: str, act: bool = False, limit: int = -1) -> None:
    conf = _load_conf(conf_fn)  # sets project_dir
    print(f">> Setting project_dir '{conf['project_dir']}'")

    wb = load_workbook(conf["excel_fn"], data_only=True)
    ws = wb[conf["sheet_title"]]  # sheet exists already

    conf["RIA"] = init_ria()  # mpApi.util.Ria's client
    init_log(act=act, conf=conf, conf_fn=conf_fn, limit=limit)
    print(f">> Getting template from RIA Object {conf['template_id']}")
    conf["templateM"] = conf["RIA"].get_template(ID=conf["template_id"], mtype="Object")

    for idx, row in enumerate(ws.iter_rows(min_row=conf["excel_row_offset"]), start=2):
        per_row(idx=idx, row=row, conf=conf, act=act)
        if limit == idx:
            print(">> Limit reached")
            break


def create_record(*, row: tuple, conf: dict, act: bool) -> None:
    # print(">> Create record")
    global hits
    hits += 1  # we're counting the records that have or would be created

    if len(conf["templateM"]) != 1:
        raise TypeError("Template does not have a single record")

    recordM = deepcopy(conf["templateM"])  # so we dont change the original template
    recordM._dropFieldsByName(element="systemField", name="__uuid")
    recordM._dropAttribs(xpath="//m:moduleItem", attrib="id")
    set_ident(
        recordM, ident=row[0].value, institution=conf["institution"]
    )  # from Excel as str
    set_ident_sort(recordM, nr=int(row[1].value))
    set_sachbegriff(recordM, sachbegriff=row[2].value)
    # problems
    set_beteiligte(recordM, beteiligte=row[3].value, conf=conf)
    set_erwerbDatum(recordM, datum=row[4].value)
    set_erwerbungsart(recordM, art=row[5].value)
    set_erwerbNr(recordM, nr=row[6].value)
    set_erwerbVon(recordM, von=row[7].value)
    set_geogrBezug(recordM, name=row[8].value)
    set_objRefA(recordM, Vorgang=row[9].value, conf=conf)
    set_invNotiz(recordM, bemerkung=row[10].value)  # Spalte L rarely filled-in

    # print(recordM)
    recordM.uploadForm()  # we need that to delete ID
    p = conf["project_dir"] / "debug.object.xml"
    print(f">> Writing record to file '{p}'")
    recordM.toFile(path=p)
    if act:
        objId = conf["RIA"].create_item(item=recordM)
        print(f">> Created record {objId} in RIA '{row[0].value}'")
    else:
        print(f">> Not creating record in RIA '{row[0].value}' (since no act)")

        # p2 = conf["project_dir"] / f"debug.object{objId}.xml"
        # print(f">> Writing to '{p2}'")
        # recordM.toFile(path=p2)


def init_log(*, act: bool, conf: dict, conf_fn: str, limit: int) -> None:
    """
    Create a simple logger at file becky20250510-0956.log
    """
    # should we only log if we actually do something with act=True?
    # to avoid plethora of log files?
    if act is True:
        now = datetime.now()
        datetime_str = now.strftime("%Y%m%d-%H%M%S")
        log_fn = f"becky{datetime_str}.log"
        logging.basicConfig(
            filename=log_fn,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
        )
        # - %(name)s is currently not necessary
        logger = logging.getLogger(__name__)
        logger.info(f"becky started with {act=} and {limit=}")
        logger.info(f"loading Excel file '{conf['excel_fn']}'")
    else:
        logger = logging.getLogger(__name__)
        logger.addHandler(logging.NullHandler())


def log_print_info(msg: str) -> None:
    """
    log and print info message simultaneously
    """
    logger = logging.getLogger(__name__)
    logger.info(msg)
    print(f"   {msg}")


def per_row(*, idx: int, row: Cell, conf: dict, act: bool) -> None:
    ident = row[0].value  # from Excel as str
    font_color = row[0].font.color
    global hits
    if font_color and font_color.rgb == "FFFF0000":  # includes the alpha channel
        print(f"***[{hits}]{idx}: {ident}")
        if m := record_exists2(ident=ident, conf=conf):
            # Wollen wir hier Fehler loggen um Nachzuvollziehen, wo die Infos aus Excel
            # nicht eingetragen wurden?
            log_print_info(f"Record '{ident}' exists already")
        else:
            create_record(row=row, conf=conf, act=act)
            if m > 1:
                logging.warning(
                    f"Multiple identNr: More than one IdentNr exists already with this number {ident}"
                )


#
# more private
#


def _load_conf(conf_fn: str) -> dict:
    print(f">> Reading configuration '{conf_fn}'")
    with open(Path(conf_fn), "rb") as toml_file:
        conf = tomllib.load(toml_file)
    conf["project_dir"] = Path(__file__).parents[4] / "sdata"  # project_dir
    return conf
