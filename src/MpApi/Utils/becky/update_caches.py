"""
We read Becky's excel, parse it for names etc., look them up in RIA and write the
results in the cache file.
"""

import argparse
from mpapi.constants import get_credentials
from mpapi.search import Search
from MpApi.Utils.becky.becky import _load_conf
from MpApi.Utils.becky.cache_ops import (
    open_archive_cache,
    open_person_cache,
    save_person_cache,
    save_archive_cache,
    set_change,
)
from MpApi.Utils.becky.set_fields_Object import (
    _sanitize_multi,
    _triple_split2,
)

from MpApi.Utils.Ria import RIA, init_ria
from openpyxl import Workbook, load_workbook, worksheet
from pathlib import Path


def main(conf_fn: str, mode: str, limit: int = -1) -> None:
    conf = _load_conf(conf_fn)

    print(f">> Reading workbook '{conf['excel_fn']}'")
    wb = load_workbook(conf["excel_fn"], read_only=True)
    ws = wb[conf["sheet_title"]]  # sheet exists already
    match mode:
        case "person":
            update_persons(conf=conf, sheet=ws, limit=limit)
        case "archive":
            update_archive(conf=conf, sheet=ws, limit=limit)
        case _:
            wb.close()
            raise SyntaxError(f"Unknown mode '{mode}'")
    wb.close()


def process_names(*, beteiligte: str, cache: dict) -> dict:
    """
    Gets called when looping through Excel, so no info from RIA yet.
    If necessary we write name, date the info in cache.
    """
    if beteiligte is None:
        return cache  # it's perfectly possible that a cell is empty...

    beteiligteL = _sanitize_multi(beteiligte)

    for count, beteiligte2 in enumerate(beteiligteL, start=1):
        name, role, date = _triple_split2(beteiligte)
        # we're counting the names in one cell here, not the lines
        # print(f"{count}:{name} [{role}]")
        # if role not in roles:
        #    roles.add(role)
        if name is None:
            continue
        if date is None:
            date = "None"
            # raise TypeError(f"Date is None! {name}")
        if name not in cache:
            print(f">> Name not yet in cache '{name}' ({date=})")
            # cache[name] = {}
            cache[name] = {date: []}
            set_change()
        else:
            if date not in cache[name]:
                print(f">> Date not yet in cache '{name}' '{date}'")
                cache[name][date] = []
                set_change()
    return cache


def query_archives(*, ident: str, client: RIA) -> list:
    """
    For a given archive ident (eg. E 1012/1898), return the id or ids.

    If there is no such record, returns empty list.
    """
    if ident is None:
        # it's purrfectly allowed for an Excel cell to be empty
        print("ident is None")
        return list()
    else:
        archive_ident = ident.strip()
        q = Search(module="Object", limit=-1, offset=0)
        q.AND()
        q.addCriterion(
            operator="equalsField",  # notEqualsTerm
            field="__orgUnit",  # __orgUnit is not allowed in Zetcom's own search.xsd
            value="EMArchiv",
        )
        q.addCriterion(
            field="ObjObjectNumberVrt",
            operator="equalsField",
            value=archive_ident,
        )
        q.addField(field="__id")
        q.validate(mode="search")  # raises if not valid
        m = client.mpapi.search2(query=q)
        return m.get_ids(mtype="Object")


def query_persons(*, name: str, date: str, client: RIA) -> list:
    """
    for a given name (Nennfom), look up the objIds in RIA and return them as a list.

    New: We expect date, include date in query and return less ids, i.e. only the
    ones that match the date.
    """
    print(f"***{name}***{date=}")
    q = Search(module="Person", limit=-1, offset=0)
    q.AND()
    q.addCriterion(
        field="PerNennformTxt",
        operator="equalsField",
        value=name,
    )
    q.addCriterion(
        field="PerDateGrp.DatingNewTxt",
        operator="equalsField",
        value=date,
    )
    q.addField(field="__id")
    q.validate(mode="search")  # raises if not valid
    m = client.mpapi.search2(query=q)
    # print(m)
    return m.get_ids(mtype="Person")


def update_archive(*, conf: dict, sheet: worksheet, limit: int) -> None:
    print(f">> Loading archive cache '{conf['archive_cache']}'")
    archive_data = open_archive_cache(conf)
    print(">> Looping thru excel looking for archival documents' idents")
    client = init_ria()
    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        print(f"Line {idx}")
        if row[9].font is None:
            print(f"...ignoring line since no color {idx}")
            continue

        font_color = row[9].font.color  # relying on red font
        if font_color and font_color.rgb == "FFFF0000":  # includes the alpha channel
            _archive_per_red_cell(row[9].value, data=archive_data, client=client)
        if idx % 500 == 0:
            save_archive_cache(data=archive_data, conf=conf)
        if limit == idx:
            save_archive_cache(data=archive_data, conf=conf)
            print(">> Limit reached")
            break
    print(f"Saving archive cache: {len(archive_data)}")
    save_archive_cache(data=archive_data, conf=conf)


def update_persons(*, conf: dict, sheet: worksheet, limit: int) -> None:
    print(f">> Loading person cache '{conf['person_cache']}'")
    person_data = open_person_cache(conf)

    person_data = _scan_excel(
        sheet=sheet, person_data=person_data, conf=conf, limit=limit
    )
    person_data = _lookup_pk_ids(person_data=person_data, conf=conf, limit=limit)
    save_person_cache(data=person_data, conf=conf)


#
# private
#
def _archive_per_red_cell(cell: str, *, data: dict, client: RIA) -> None:
    # may contain multiple values separated by ;
    if cell is not None:
        identL = cell.split(";")  #
        identL = [element.strip() for element in identL]

        for ident in identL:
            print(f"***{cell} -> {ident=}")
            if ident in data:
                if len(data[ident]) == 0:
                    # Do we want to re-check empty lists?
                    # ident exists, but list is empty
                    # _query_archives(ident, client, data)
                    pass
            else:  # ident not (yet) in cache
                if ident is None:
                    data[ident] = list()
                else:
                    print(f">> querying archives '{ident}'")
                    idL = query_archives(ident=ident, client=client)
                    set_change()
                    print(f"{idL=}")
                    data[ident] = idL  # may be empty list


def _lookup_pk_ids(*, person_data: dict, conf: dict, limit: int) -> dict:
    client = init_ria()
    print(">> Unidentified names?")
    for idx, name in enumerate(person_data, start=1):
        for date in person_data[name]:
            # print (f"**{date=}")
            if not person_data[name][date]:  # if tuple is empty
                # where do we get date from? _each_person
                idL = query_persons(client=client, name=name, date=date)
                person_data[name][date] = idL
                set_change()
                print(f"{idL}")
            if idx % 25 == 0:
                save_person_cache(data=person_data, conf=conf)
            if limit == idx:
                print(">> Limit reached")
                break
    return person_data


def _scan_excel(*, sheet: worksheet, person_data: dict, conf: dict, limit: int) -> dict:
    """
    Many rows repeat the same name, so we first make an index with distinct entries.
    """
    print(">> Looping thru excel looking for names")
    for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if idx % 100 == 0:
            print(f"Line {idx} -- {len(person_data)} distinct names found")
        if row[0].font is None:
            print(f"...ignoring line since no color {idx}")
            continue
        font_color = row[0].font.color
        if font_color and font_color.rgb == "FFFF0000":  # includes the alpha channel
            person_data = process_names(beteiligte=row[3].value, cache=person_data)
        if idx % 1000 == 0:
            print("Saving cache")
            save_person_cache(data=person_data, conf=conf)
        if limit == idx:
            print(">> Limit reached")
            break
    # set_change()
    print("Saving cache")
    save_person_cache(data=person_data, conf=conf)
    print(f"Initial excel scan done: {len(person_data)} distinct names found")
    return person_data


#
# script
#


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("conf", help="Location of becky_conf.toml")
    parser.add_argument(
        "-l",
        "--limit",
        help="Stop after a number of rows in Excel file are processed.",
        type=int,
    )
    parser.add_argument(
        "-m",
        "--mode",
        type=str,
        help="Pick which cache(s) to update.",
        choices=["person", "archive"],
    )
    args = parser.parse_args()
    main(conf_fn=args.conf, limit=args.limit, mode=args.mode)
