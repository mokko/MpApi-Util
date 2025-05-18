"""

Composition over inheritance. So we're moving Excel related stuff from BaseApp.py to this class

USAGE
    xls = Xls(path="test.xlsx")
    xls.save()
    xls.backup()
    wb = xls.get_or_create_wb()
"""

from openpyxl import Workbook, load_workbook  # worksheet
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import shutil
import sys
from tqdm import tqdm
from typing import Any, Iterator

red = Font(color="FF0000")
teal = Font(color="008080")
blue = Font(color="0000FF")


class ConfigError(Exception):
    pass


class NoContentError(Exception):
    pass


class Xls:
    def __init__(self, path: str | Path, description: dict[str, str]) -> None:
        """
        currently there can be only one description.
        """
        self.path = Path(path)
        self.backup_fn = Path(str(self.path) + ".bak")  # ugly
        self.wb = self.get_or_create_wb()
        self.shutdown_requested = False
        self.changed = False  # keep a state to know if saving is necessary
        self.description = description

    def backup(self) -> bool:
        """
        Write a new backup file of the Excel file.
        """
        try:
            shutil.copy(self.path, self.backup_fn)
        except KeyboardInterrupt:
            self.request_shutdown()
        return True

    def backup_if_change(self) -> bool:
        if self.changed:
            return self.backup()

    def set_change(self) -> None:
        """
        Set the object variable changed to signal that save is necessary.
        """
        self.changed = True

    def drop_row_if_file_gone(self, *, col: str = "A", sheet: Worksheet) -> None:
        """
        Loop thru Excel sheet "Assets" and check if the files still exist. We use
        relative filename for that, so update has to be executed in right dir.
        If the file no longer exists on disk (e.g. because it has been renamed),
        we delete it from the excel sheet by deleting the row.

        This is for the scandir step. NOT USED AT THE MOMENT.
        """
        print("Checking for file changes")
        c = 3
        with tqdm(total=sheet.max_row - c) as pbar:
            for row in sheet.iter_rows(min_row=c):  # start at 3rd row
                filename = sheet[f"{col}{c}"].value
                pbar.update()
                if filename is not None:
                    if not Path(filename).exists():
                        print(f"Deleting Excel row {c} file gone '{filename}'")
                        sheet.delete_rows(c)
                c += 1
        print("   done")

    def file_exists(self) -> bool:
        """
        Returns True if Excel Excel file (at self.path) exists at specified location.
        """
        return self.path.exists()

    def get_conf(self, *, cell: str, default: Any = None) -> str | None:
        """
        Returns the specified field from sheet conf or None if field is empty or
        only consists of space.
        """
        conf_ws = self.wb["Conf"]
        value = conf_ws[cell].value  # can be None
        if value is None:
            value = default
        elif isinstance(value, str):
            if value.isspace() or value == "":
                value = default
        return value

    def get_conf2(self, *, cell: str, default: Any = None) -> str:
        """
        Like get_conf, but returns empty string instead of None.
        """
        ret = self.get_conf(cell=cell, default=default)
        if ret is None:
            return ""
        return ret

    def get_conf_required(self, *, cell: str, default: Any = None) -> str:
        """
        Like get_conf, but raises if return value is None.
        """
        ret = self.get_conf(cell=cell, default=default)
        if ret is None:
            raise ConfigError(f"ERROR: Config value {cell} missing")
        return ret

    def get_conf_true(self, *, cell: str) -> bool:
        """
        Version of get_conf that returns True if Excel cell has str "true" (case
        insensitive) or False if something ele.
        """
        cell = self.get_conf(cell=cell)
        if cell is None or cell.isspace() or cell == "":
            return False
        if cell.lower() == "true":
            return True
        else:
            return False

    def get_sheet(self, *, title: str) -> Worksheet:
        try:
            ws = self.wb[title]
        except:
            raise ConfigError(f"ERROR: Excel file has no sheet {title}")
        return ws

    def get_or_create_sheet(self, *, title: str) -> Worksheet:
        try:
            ws = self.wb[title]  # sheet exists already
        except:  # new sheet
            ws = self.wb.active
            if ws.title == "Sheet":
                ws.title = title
            else:
                self.wb.create_sheet(title)
            ws = self.wb[title]
            self.changed = True
        return ws

    def get_or_create_wb(self) -> Workbook:
        """
        Given a file path for an excel file in self.path, return the respective workbook
        or make a new one if the file doesn't exist. We also save wb internally.
        """
        try:
            return self.wb
        except:
            if self.path.exists():
                # print (f"* Loading existing excel: '{data_fn}'")
                self.wb = load_workbook(self.path, data_only=True)
                return self.wb
            else:
                # print (f"* Starting new excel: '{data_fn}'")
                self.changed = True
                self.wb = Workbook()
                return self.wb

    def load_workbook(self) -> Workbook:
        """
        Load workbook from self.path or raise if file does not exist. Sets self.wb and
        returns workbook.
        """
        if not self.path.exists():
            raise ConfigError(f"ERROR: Excel file not found! {self.path}")

        self.wb = load_workbook(self.path, data_only=True)
        return self.wb

    def loop(
        self,
        *,
        sheet: Worksheet,
        offset: int = 3,
        limit: int = -1,
    ) -> Iterator:
        """
        Loop thru the rows of specified sheet.

        Returns a dict of cells in dict as well as the current row number (rno):
        for c,rno in self.loop(sheet=ws, limit=self.limit):
            print (f"row number {rno} {c['filename']}")

        For this to work, we need a description (self.description).
        """
        for rno, row in enumerate(
            sheet.iter_rows(min_row=offset), start=offset
        ):  # start at 3rd row
            cells = self._rno2dict(rno, sheet)
            yield cells, rno
            if limit == rno:
                print("* Limit reached")
                break

    def loop2(
        self,
        *,
        sheet: Worksheet,
        offset: int = 3,
        limit: int = -1,
    ) -> Iterator:
        """
        Loop thru the rows of specified sheet.

        A version that returns the row and doesn't require description dictionary:
            for row,rno in self.loop2(sheet=ws, limit=self.limit):
                print (f"row number {rno} {row[0]}")
        """

        for rno, row in enumerate(
            sheet.iter_rows(min_row=offset), start=offset
        ):  # start at 3rd row
            yield row, rno
            if limit == rno:
                print("* Limit reached")
                break

    def make_conf(self, conf: dict[str, str]) -> None:
        """
        Create a sheet named Conf if it doesn't exist yet and fill it with values from
        the dict conf. Also do some formatting (first row bold etc.).
        """
        conf_ws = self.get_or_create_sheet(title="Conf")
        max_row = 0
        for cell_label in conf:
            conf_ws[cell_label] = conf[cell_label]
            no = int(cell_label[-1])
            if no > max_row:
                no = max_row

        for col in ["A", "B", "C"]:
            conf_ws.column_dimensions[col].width = 25

        for row in conf_ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)

        for row in conf_ws.iter_rows(min_col=3, max_col=3):
            for cell in row:
                cell.font = blue
        self.changed = True

    def path_exists(
        self,
        *,
        cno: int = 0,
        path: Path | str,
        sheet: Worksheet,
    ) -> int | None:
        """
        Returns row number as int if filename is already in Excel sheet at specified row,
        else None.

        We are using the first column as default, but can also be specified (cno).

        Note that when you are using relative path, you need to be in correct directory.

        We used to use the filename from column A which SHOULD be unique in the
        MuseumPlus context, but which is strange requirement in the world of directories,
        where multiple dirs may contain files with the same name.

        Now we use a cno (no 8) column as int.
        """

        for idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
            # start at 3rd row
            fn = row[cno].value
            # print (f"_path_in_list: {fn=}{name=}")
            if fn == str(path):
                print(f"WARN: Known full path '{path}' (not adding to list)")
                return idx
        return None

    def raise_if_conf_value_missing(self, required: dict) -> None:
        base_msg = "ERROR: Missing configuration value: "
        conf_ws = self.wb["Conf"]
        for cell in required:
            if conf_ws[cell].value is None:
                raise ConfigError(base_msg + required[cell])

    def raise_if_content(self, sheet: Worksheet) -> bool:
        """
        Raises if sheet has more than 2 lines.
        """
        if sheet.max_row > 2:
            raise NoContentError(f"ERROR: Excel contains {sheet.max_row} rows!")
        return False

    def raise_if_file(self) -> bool:
        """
        Raise if file exists already; returns False if file does NOT exist.
        """
        if self.path.exists():
            raise Exception(f"ERROR: {self.path} exists already!")
        return False

    def raise_if_not_initialized(self, sheet: Worksheet) -> None:
        if sheet.max_row < 2:
            raise ConfigError("ERROR: Excel file not initialized!")

    def raise_if_no_content(self, sheet: Worksheet) -> bool:
        """
        Assuming that after init excel has to have more than 2 lines.
        Returns False if there is content.
        """

        if sheet.max_row < 3:
            raise NoContentError(
                f"ERROR: no data found; excel contains {sheet.max_row} rows!"
            )
        return False

    def raise_if_no_file(self) -> bool:
        """
        Raise if no file at self.path.
        """

        if not self.path.exists():
            raise Exception(f"ERROR: {self.path} does NOT exist!")
        return False

    def real_max_row(self, sheet: Worksheet) -> int:
        """
        I might need to find the last row that has content if ws.max_row is not realiable
        and often includes empty rows.

        In that case I would take max_row and go back to find the last row with content
        in cell A.
        """
        real_max = sheet.max_row
        while sheet[f"A{real_max}"].value in (None, ""):
            real_max -= 1
        return real_max

    def request_shutdown(self):
        """
        Prints a message and changes class variable. To be called in except KeyboardInterrupt.
        """
        print("Keyboard interrupt recieved, requesting shutdown...")
        self.shutdown_requested = True

    def save(self) -> bool:
        """
        Made this only to have same print msgs all the time
        """
        print(f"   saving {self.path}")
        try:
            self.wb.save(filename=self.path)
        except KeyboardInterrupt:
            self.request_shutdown()
        else:
            self.changed = False
        return True

    def save_bak_shutdown(
        self, *, rno: int | None, save: int = 1, bak: int = 10
    ) -> None:
        """
        shutdown if requested, make a backup every n times (bak) and save
        current status every m times (save).
        """
        self.shutdown_if_requested()
        if rno is not None and rno % bak == 0:
            self.backup()
        if rno is not None and rno % save == 0:
            self.save_if_change()

    def save_if_change(self) -> bool:
        """
        Version of save that saves only if changes were registered in variable
        self.changed.
        """
        if self.changed:
            self.save()
            return True
        return False

    def save_and_shutdown_if_requested(self) -> None:
        self.save()
        self.changed = False
        if self.shutdown_requested:
            print("Planned shutdown.")
            sys.exit(0)

    def shutdown_if_requested(self) -> None:
        """
        Do the shutdown if class variable is set. To be used in the loop at an appropriate time.

        To be sure, we include a save here. That is the usual order should be
        self.shutdown_if_requested()
        self.save()
        We could also rename this to save_and_shutdown_if_requested() and save one line.
        """
        if self.shutdown_requested:
            self.save()
            print("Planned shutdown.")
            sys.exit(0)

    def wipe(self, *, sheet: Worksheet) -> None:
        """Delete everything but the header. Slow."""
        rno = 3
        with tqdm(total=sheet.max_row - 2) as pbar:
            while rno <= sheet.max_row:
                # print(f"wiping row {rno}")
                pbar.update()
                sheet.delete_rows(rno)
        self.changed = True
        self.save()

    def write_header(self, *, sheet: Worksheet) -> None:
        """
        Take the table description (a dict) and write it to the top two lines of the
        specified worksheet.

        The table description is a dictionary that is structured as follows:
        self.description = {
            "filename": { # short label
                "label": "Asset Dateiname",
                "desc": "aus Verzeichnis",
                "col": "A",
                "width": 20,
            },
        }
        """

        for short in self.description:
            col = self.description[short]["col"]  # single letter
            sheet[f"{col}1"] = self.description[short]["label"]
            sheet[f"{col}1"].font = Font(bold=True)
            if "desc" in self.description[short]:
                desc_txt = self.description[short]["desc"]
                sheet[f"{col}2"] = desc_txt
                sheet[f"{col}2"].font = Font(size=9, italic=True)
            if "width" in self.description[short]:
                width = self.description[short]["width"]
                sheet.column_dimensions[col].width = width

    #
    # private
    #

    def _rno2dict(self, rno: int, sheet: Worksheet) -> dict[str, Any]:
        """
        We read  the provide a dict with labels as keys based on table description
        (self.description).
        """
        cells = dict()
        for label in self.description:
            col = self.description[label]["col"]
            cells[label] = sheet[f"{col}{rno}"]
        return cells
