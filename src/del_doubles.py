"""
We made a mistake and created too many assets (by creating and attaching the same resource multiple times).

Now we want to fix the mistake, but looping thru the Excel file and deleting the additional assets.

That is we scan the Excel table and look for assets.
"""

from openpyxl import load_workbook
from mpapi.constants import get_credentials
from mpapi.client import MpApi

limit = -1

xls_fn = "upload14.xlsx"


def del_double():
    wb = load_workbook(xls_fn)
    print("Saving excel")
    wb.save(xls_fn)
    ws = wb["Assets"]
    user, pw, baseURL = get_credentials()
    ria = MpApi(baseURL=baseURL, user=user, pw=pw)
    print(f"Logging in with user {user}")
    c = 3  # counter; used report different number
    for row in ws.iter_rows(min_row=c):  # start at 3rd row
        per_row(ria, row, c)
        if limit == c:
            print("limit reached!")
            break
        c += 1
    print("Saving excel")
    wb.save(xls_fn)
    print("done!")


def per_row(ria, row, c):
    assets = row[2].value.split(";")
    if assets[1:]:
        for mulId in assets[1:]:
            print(f"{c} Del Multimedia {mulId}...")
            r = ria.deleteItem2(mtype="Multimedia", ID=int(mulId))
            print(r)
    row[2].value = assets[0]
    # print(f"::{assets[1:]}")


if __name__ == "__main__":
    del_double()
