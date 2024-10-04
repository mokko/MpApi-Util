"""
We read Becky's excel, parse it for names, look them up in RIA and write the
results in the cache file.
"""

from becky import _load_conf
from mpapi.constants import get_credentials
from mpapi.search import Search
from MpApi.Utils.person_cache import open_cache, save_cache
from MpApi.Utils.set_fields_Object import _each_person
from MpApi.Utils.Ria import RIA, init_ria
from openpyxl import Workbook, load_workbook, worksheet
from pathlib import Path

conf_fn = "becky_conf.toml"  # in sdata
# roles = set()


def main() -> None:
    conf = _load_conf()
    person_data = {}

    print(">> Reading workbook")
    wb = load_workbook(conf["excel_fn"], data_only=True)
    ws = wb[conf["sheet_title"]]  # sheet exists already

    print(f">> Loading person cache '{conf["person_cache"]}'")
    person_data = open_cache(conf)

    print(">> Looping thru table")
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # print(f"Line {idx}")
        person_data = process_names(beteiligte=row[3].value, cache=person_data)

    save_cache(data=person_data, conf=conf)

    print(">> Unidentified names?")
    client = init_ria()
    for idx, name in enumerate(person_data, start=1):
        if not person_data[name]:  # if tuple is empty
            idL = query(client=client, name=name)
            person_data[name] = idL
            print(idL)
        # if idx % 25 == 0:
        #    save_cache(data=person_data, conf=conf)
    save_cache(data=person_data, conf=conf)
    # print(roles)


def process_names(*, beteiligte: str, cache: dict) -> dict:
    if beteiligte is None:
        return cache  # it's perfectly possible that a cell is empty...
    for count, (name, role) in enumerate(_each_person(beteiligte), start=1):
        # we're counting the names in one cell here, not the lines
        # print(f"{count}:{name} [{role}]")
        # if role not in roles:
        #    roles.add(role)
        if name not in cache:
            cache[name] = []
    return cache


def query(*, name: str, client: RIA) -> list:
    result = tuple()
    q = Search(module="Person", limit=-1, offset=0)
    # q.AND()
    q.addCriterion(
        field="PerNennformTxt",
        operator="equalsField",
        value=name,
    )
    q.addField(field="__id")
    q.validate(mode="search")  # raises if not valid
    m = client.mpapi.search2(query=q)
    m.get_ids(mtype="Person")
    # print(m)
    return m.get_ids(mtype="Person")


#
# script
#


if __name__ == "__main__":
    main()
