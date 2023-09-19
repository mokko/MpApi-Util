from openpyxl import Workbook, load_workbook
from pathlib import Path
import sys

path = Path(sys.argv[1])
fn_new = "excel.xlsx"


print(f"working on {path}")

wb = load_workbook(path, data_only=True)
ws = wb["Assets"]

c = 2  # counter; assuming it's zero-based
deleted_rows = 0
while c < ws.max_row:
    fn = ws.cell(row=c, column=1).value
    if fn is not None:
        print(f"{c}:{fn}")
        if fn.startswith("I_MV_") and "__" not in fn:
            print(f"   rm this row")
            ws[f"C{c}"] = "x"
            ws.delete_rows(c)
            c -= 1
            deleted_rows += 1
    c += 1  # only add 1 to line counter if we didn't delete the row

print(f"Writing to {fn_new} total deleted row {deleted_rows}")
try:
    wb.save(filename=fn_new)
except KeyboardInterrupt:
    print("Catching keyboard interrupt during Excel operation; try again...")
