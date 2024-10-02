"""
Script to import records from Excel for Kamerun project

Fields that have to fill in
- IdentNr
- Sachbegriff
- Beteiligte + Rolle
- Erwerb.Datum
- Erwerbungsart
- Erwerb. Nr.
- Erwerbung von
- Geogr. Bezug
- Obj. Referenz A
fraglich
- Obj. Referenz B
- Inventarnotiz

Todo:
- Why dont we need to provide a namespace?
- Put the template record into a group ("Kamerun 2024"), so all created records are in a group
"""

import argparse
from copy import deepcopy
from lxml import etree  # t y p e : i g n o r e

from mpapi.constants import get_credentials
from mpapi.module import Module
from mpapi.search import Search

from MpApi.Utils.Ria import RIA, init_ria
from MpApi.Utils.set_fields_Object import (
    set_ident,
    set_ident_sort,
    set_sachbegriff,
    set_beteiligte,
    set_erwerbDatum,
    set_erwerbungsart,
    set_erwerbNr,
    set_erwerbVon,
    set_geogrBezug,
    set_objRefA,
)
from MpApi.Utils.Xls import Xls

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
from pathlib import Path

#
# CONFIGURATION
#

conf = {
    "excel_fn": Path(
        r"C:\Users\M-MM0002\Work Folders\Eigene Dateien\Kamerun\Bis 10224 Abschrift_HK_Afrika_III_C.xlsx"
    ),
    "excel_row_offset": 2,  # first line that needs to be processed, 1-based
    "institution": "EM",
    "org_unit": "EMAfrika1",
    "sheet_title": "Sheet1",
    "template_id": 3659476,  # Object
}


def create_record(*, row: tuple, template: Module, client: RIA, act: bool) -> None:
    # print(">> Create record")

    if len(template) != 1:
        raise TypeError("Template does not have a single record")

    recordM = deepcopy(template)  # so we dont change the original template
    recordM
    set_ident(
        recordM, ident=row[0].value, institution=conf["institution"]
    )  # from Excel as str
    set_ident_sort(recordM, nr=int(row[1].value))
    set_sachbegriff(recordM, sachbegriff=row[2].value)
    set_beteiligte(recordM, beteiligte=row[3].value)
    set_erwerbDatum(recordM, datum=row[4].value)
    set_erwerbungsart(recordM, art=row[5].value)
    set_erwerbNr(recordM, nr=row[6].value)
    set_erwerbVon(recordM, von=row[7].value)
    set_geogrBezug(recordM, name=row[8].value)
    set_objRefA(recordM, keineAhnung=row[9].value)

    # print(recordM)
    recordM.uploadForm()  # we need that to delete ID
    recordM.toFile(path=f"../sdata/debug.object.xml")
    if act:
        objId = client.create_item(item=recordM)
        print(f">> Created record {objId} in RIA ({row[0].value}")
        recordM.toFile(path=f"../sdata/debug.object{objId}.xml")


def main(*, limit: int = -1, act: bool = False) -> None:
    wb = load_workbook(conf["excel_fn"], data_only=True)
    ws = wb[conf["sheet_title"]]  # sheet exists already

    client = init_ria()

    print(f">> Getting template from RIA Object {conf['template_id']}")
    tmplM = client.get_template(ID=conf["template_id"], mtype="Object")

    for idx, row in enumerate(ws.iter_rows(min_row=conf["excel_row_offset"]), start=2):
        per_row(idx=idx, row=row, template=tmplM, client=client, act=act)
        if limit == idx:
            print(">> Limit reached")
            break


def per_row(*, idx: int, row: Cell, template: Module, client: RIA, act: bool) -> None:
    ident = row[0].value  # from Excel as str
    font_color = row[0].font.color
    if font_color and font_color.rgb == "FFFF0000":  # includes the alpha channel
        print(f"{idx}: {ident} red")
        if record_exists(ident=ident, client=client):
            print(f"   Record '{ident}' exists already")
        else:
            create_record(template=template, row=row, client=client, act=act)


def record_exists(*, ident: str, client: RIA) -> bool:
    """
    Check ria if a record with a specific identNr exists. It's not relevant if one or
    multiple results exist.
    N.B. This search is not exact.
    """
    q = Search(module="Object", limit=-1, offset=0)
    q.AND()
    q.addCriterion(
        field="ObjObjectNumberVrt",
        operator="equalsField",
        value=str(ident),
    )
    q.addCriterion(field="__orgUnit", operator="equalsField", value=conf["org_unit"])
    q.addField(field="__id")
    q.validate(mode="search")  # raises if not valid
    m = client.mpapi.search2(query=q)
    if len(m) > 1:
        raise TypeError("Warning! more than one result in record_exists")
    if m:
        return True
    else:
        return False


#
# script
#


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-a", "--act", help="Actually change RIA", action="store_true")
    parser.add_argument(
        "-l",
        "--limit",
        help="Stop after a number of rows in Excel file are processed.",
        type=int,
    )
    args = parser.parse_args()
    main(limit=args.limit, act=args.act)
