"""
We create a taylor-made Excel file for UK. Let's keep it simple and execute it
from the individual folders multiple times manually.

excel_fn comes from the dirname and lists filenames in column 1.

- Where do we start with the program? 
- How do we call the sheets?
- How many Excel sheets do we make?

Perhaps it's easier if we have to execute

"""

from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Font  # Alignment,
from pathlib import Path
import sys

# cd manually to the dir you want
# src_dir = Path(r"\\pk.de\smb\Mediadaten\Projekte\AKU\MDVOS-Bildmaterial\2023-05-00_FINAL_EM_VisAnthro_Dias_Konrad") # \001-00.602-00.0800


def add_file(ws: worksheet, rno: int, fn: Path) -> None:
    """
    Receive a worksheet and a file path. List it in the Excel sheet.

    At this point, we dont check if the filename been entered before, so potentially we
    could enter it multiple times.
    """

    # print(f"*{rno}: {fn.name}")
    ws[f"A{rno}"] = fn.name


def mk_excel(fn: Path) -> (Workbook, worksheet):
    """
    Creates a new excel file in memory.

    When do we save it to disk?
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dias"
    ws["A1"] = "Dateiname"
    ws.column_dimensions["A"].width = 20
    ws["B1"] = "Foto Nr"
    ws.column_dimensions["B"].width = 15
    ws["C1"] = "Datum/Jahr Aufnahme"
    ws.column_dimensions["C"].width = 10
    ws["D1"] = "Ort Aufnahme"
    ws.column_dimensions["D"].width = 20
    ws["E1"] = "Beschreibung/Stichwörter"
    ws.column_dimensions["E"].width = 40
    ws["F1"] = "Sonstiges"
    ws.column_dimensions["F"].width = 40

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    return wb, ws


def save(wb: Workbook, fn: Path) -> None:
    wb.save(filename=fn)


limit = -1
src_dir = Path().cwd()
excel_fn = Path(src_dir.name + ".xlsx")
if excel_fn.exists():
    print(f"Excel exists already '{excel_fn}'")
    sys.exit(0)
else:
    wb, ws = mk_excel(excel_fn)
    c = 1  # we count files
    for p in src_dir.glob("**/*.tif"):
        print(f"{c}: {p.name}")
        add_file(ws, c + 1, p)
        if limit and c == limit:
            print("Limit reached!")
            save(wb, excel_fn)
            break
        c += 1
    save(wb, excel_fn)
